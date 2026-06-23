#!/usr/bin/env python3
"""Generate docs/Page_Views_List.xlsx from the authoritative page view inventory.

Data sourced from docs/Page_Views_List.md (generated 2026-06-10 from pageRegistry.ts).
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

# Output path
OUTPUT = "docs/Page_Views_List.xlsx"

# Professional styles
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
TITLE_FONT = Font(name="Arial", size=16, bold=True)
SUBTITLE_FONT = Font(name="Arial", size=10, italic=True, color="666666")
NORMAL_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", size=10, bold=True)

READ_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # light green
NONE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # light red
ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin', color='B4B4B4'),
    right=Side(style='thin', color='B4B4B4'),
    top=Side(style='thin', color='B4B4B4'),
    bottom=Side(style='thin', color='B4B4B4'),
)

CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Data: Component Groups (11)
COMPONENT_GROUPS = [
    ["cmp-dashboard", "Dashboard / Command Center", "read", 10],
    ["cmp-policy-library", "Policy Library", "read", 20],
    ["cmp-forms", "Forms", "read", 30],
    ["cmp-ces", "CES / Compliance Execution", "read", 40],
    ["cmp-calendar", "Calendar", "read", 50],
    ["cmp-evidence", "Evidence Center", "read", 60],
    ["cmp-audit", "Audit Mode", "read", 70],
    ["cmp-journey", "Journey / Training", "read", 80],
    ["cmp-staffing", "Staffing / Clinical", "read", 90],
    ["cmp-iadministrator", "iAdministrator", "read", 100],
    ["cmp-user-management", "User Management / Identity Admin", "none", 110],
    ["cmp-system", "System / Settings", "read", 120],
]

# Data: Page Views Master Table (PAGE_REGISTRY entries)
PAGE_VIEWS_MASTER = [
    ["page.dashboard", "Dashboard", "/dashboard", "cmp-dashboard", "read", "dashboard.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE] [NAV]", "Yes (Primary)"],
    ["page.library", "Policy Library", "/library", "cmp-policy-library", "read", "policyLibrary.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE] [NAV]", "Yes (via Taxonomy)"],
    ["page.policy-detail", "Policy Detail", "/library/:policyId", "cmp-policy-library", "read", "policyLibrary.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE]", "Detail only"],
    ["page.policy-lifecycle", "Policy Lifecycle", "/policy-lifecycle", "cmp-policy-library", "read", "policyLifecycle.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE] [NAV]", "Yes (Direct)"],
    ["page.framework", "Framework", "/framework", "cmp-policy-library", "read", "frameworkTaxonomy.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE] [NAV]", "Yes (via Taxonomy)"],
    ["page.taxonomy", "Taxonomy", "/taxonomy", "cmp-policy-library", "read", "frameworkTaxonomy.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE] [NAV]", "Yes (Direct)"],
    ["page.achc-survey", "ACHC Survey Alignment", "/framework/achc-survey", "cmp-policy-library", "read", "surveyor.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE]", "No"],
    ["page.forms", "Forms Library", "/forms", "cmp-forms", "read", "forms.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE] [NAV]", "Yes (via Taxonomy)"],
    ["page.form-viewer", "Form Viewer / Sign", "/forms/:formId", "cmp-forms", "read", "ecign.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE]", "Detail only"],
    ["page.ces-calendar", "CES Calendar", "/ces/calendar", "cmp-ces", "read", "ces.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE] [NAV]", "Yes (CES sub)"],
    ["page.ces-board", "CES Sprint Board", "/ces/board", "cmp-ces", "read", "ces.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE] [NAV]", "Yes (CES sub)"],
    ["page.ces-workloads", "CES Workloads", "/ces/workloads", "cmp-ces", "read", "ces.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE] [NAV]", "Yes (CES sub)"],
    ["page.ces-reports", "CES Reports", "/ces/reports", "cmp-ces", "read", "ces.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE] [NAV]", "Yes (CES sub)"],
    ["page.my-tasks", "My Tasks", "/my-tasks", "cmp-ces", "read", "pmTasks.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE] [NAV]", "Yes (Direct + Mobile)"],
    ["page.workflows", "Workflows Library", "/workflows", "cmp-ces", "read", "workflows.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE] [NAV]", "Yes (CES sub)"],
    ["page.master-controls", "Master Control Inventory", "/compliance/master-controls", "cmp-ces", "read", "masterControlInventory.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE] [NAV]", "Yes (CES sub)"],
    ["page.pm-tasks", "PM — My Tasks", "/pm/my-tasks", "cmp-ces", "read", "pmTasks.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE] [NAV]", "Yes (PM nav)"],
    ["page.pm-sprint-plan", "PM — Sprint Plan", "/pm/sprint-plan", "cmp-ces", "read", "pmTasks.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE] [NAV]", "Yes (PM nav)"],
    ["page.pm-sprint-review", "PM — Sprint Review", "/pm/sprint-review", "cmp-ces", "read", "pmTasks.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE] [NAV]", "Yes (PM nav)"],
    ["page.pm-approvals", "PM — Approvals", "/pm/approvals", "cmp-ces", "read", "pmTasks.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE] [NAV]", "Yes (PM nav)"],
    ["page.pm-dashboard", "PM — Dashboard", "/pm/dashboard", "cmp-ces", "read", "pmTasks.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE] [NAV]", "Yes (PM nav)"],
    ["page.calendar", "Master Calendar", "/calendar", "cmp-calendar", "read", "calendar.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE] [NAV]", "Yes (Mobile + indirect)"],
    ["page.evidence", "Evidence Center", "/evidence", "cmp-evidence", "read", "evidence.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE] [NAV]", "Yes (Direct + CES)"],
    ["page.audit", "Audit Mode", "/audit", "cmp-audit", "read", "audit.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE] [NAV]", "Yes (CES sub)"],
    ["page.journey-home", "Journey Home", "/journey", "cmp-journey", "read", "journey.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE] [NAV]", "Yes (Direct)"],
    ["page.journey-v1", "Journey v1", "/journey/v1-journey", "cmp-journey", "read", "journey.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE] [NAV]", "Yes (Onboarding sub)"],
    ["page.journey-appendix-f", "Journey — Appendix F", "/journey/appendix-f", "cmp-journey", "read", "journey.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE] [NAV]", "Yes (Onboarding sub)"],
    ["page.journey-module", "Journey Module Player", "/journey/module/:moduleId", "cmp-journey", "read", "journey.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE]", "Detail only"],
    ["page.journey-supervisor", "Journey Supervisor View", "/journey/supervisor", "cmp-journey", "read", "journey.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE] [NAV]", "Yes (Onboarding sub)"],
    ["page.journey-admin", "Journey Admin", "/journey/admin", "cmp-journey", "read", "journey.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE] [NAV]", "Yes (Onboarding sub)"],
    ["page.journey-guide", "Journey User Guide", "/journey/guide", "cmp-journey", "read", "journey.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE] [NAV]", "Yes (Onboarding sub)"],
    ["page.onboarding-v2", "Onboarding v2", "/onboarding-v2", "cmp-journey", "read", "onboardingV2.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE] [NAV]", "Yes (Direct)"],
    ["page.clinicians", "Clinician Profiles", "/clinicians", "cmp-staffing", "read", "clinicians.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE] [NAV]", "Yes (Direct)"],
    ["page.clinician-detail", "Clinician Detail", "/clinicians/:clinicianId", "cmp-staffing", "read", "clinicians.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE]", "Detail only"],
    ["page.patients", "Patient Profiles", "/patients", "cmp-staffing", "read", "patients.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE] [NAV]", "Yes (Direct)"],
    ["page.patient-detail", "Patient Detail", "/patients/:patientId", "cmp-staffing", "read", "patients.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE]", "Detail only"],
    ["page.staffing-calendar", "Staffing Calendar", "/staffing-calendar", "cmp-staffing", "read", "staffing.calendar.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE] [NAV]", "Yes (Direct)"],
    ["page.iadministrator", "iAdministrator (Brad)", "/iadministrator", "cmp-iadministrator", "read", "brad.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE] [NAV]", "Yes (Direct)"],
    ["page.user-assignments", "User Assignments", "/admin/users", "cmp-user-management", "none", "admin.users.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE] [NAV]", "Yes (Admin sub)"],
    ["page.user-groups", "User Groups", "/admin/user-groups", "cmp-user-management", "none", "admin.userGroups.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE] [NAV]", "Yes (Admin sub)"],
    ["page.admin-roles", "Roles", "/admin/roles", "cmp-user-management", "none", "admin.roles.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE] [NAV]", "Yes (Admin sub)"],
    ["page.admin-permissions", "Permissions", "/admin/permissions", "cmp-user-management", "none", "admin.permissions.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE] [NAV]", "Yes (Admin sub)"],
    ["page.page-access", "Page View Access", "/admin/users#page-access", "cmp-user-management", "none", "(none)", "[PAGE_REGISTRY] [ROUTE]", "Admin only"],
    ["page.help-center", "Help Center", "/help", "cmp-system", "read", "helpCenter.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE] [NAV]", "Yes (Direct)"],
    ["page.system-documentation", "System Documentation", "/system-documentation", "cmp-system", "read", "systemDocumentation.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE] [NAV]", "Yes (Direct + subs)"],
    ["page.demo", "Demo Page", "/demo", "cmp-system", "read", "demo.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE] [NAV]", "Yes (Direct)"],
    ["page.hubstaff", "Hubstaff Staging", "/hubstaff", "cmp-system", "read", "hubstaff.view", "[PAGE_REGISTRY] [FEATURE] [ROUTE] [NAV]", "Yes (Direct)"],
]

# Data: Additional Surfaces (not individually in PAGE_REGISTRY)
ADDITIONAL_SURFACES = [
    ["mobileIncidentExecution.view", "Mobile Incident Execution (all states)", "/calendar/event/:eventId/*", "mobileIncidentExecution.view", "[FEATURE] [ROUTE]", "Partial (via /calendar)", "Multiple sub-stages"],
    ["bradProposal.view", "Brad Proposal", "/brad-proposal", "bradProposal.view", "[FEATURE] [ROUTE]", "No (hidden)", "Executive-only"],
    ["governance", "Governance Page", "/governance", "(RoleGate)", "[ROUTE]", "No", "RoleGate only"],
    ["artifacts / viewer", "Artifact & Generic Viewers", "/artifacts/:id, /viewer/:ref, /events/:, /tasks/:", "—", "[ROUTE]", "No", "Detail viewers"],
    ["print", "Print Views", "/print/* , /forms/:id/print , /surveyor/policy/:id", "—", "[ROUTE]", "No", "Standalone (outside shell)"],
    ["ui-staging*", "Visual Lab / Staging", "/ui-staging*", "—", "[ROUTE]", "No", "Dev/staging only"],
    ["onboarding-v2/* (nested)", "Onboarding v2 children", "/onboarding-v2/*", "onboardingV2.view", "[FEATURE] [ROUTE]", "Yes (parent)", "Covered by page.onboarding-v2"],
    ["journey/staging/m01", "Staging M01 (env-gated)", "/journey/staging/m01", "journey.view", "[FEATURE] [ROUTE]", "Conditional", "VITE_STAGING_M01 only"],
    ["admin redirects", "Admin alias redirects", "/admin , /security/identity/*", "various admin.*", "[ROUTE]", "N/A", "Redirects only"],
]

def apply_header_style(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER

def apply_cell_style(cell, is_alt=False, access_value=None):
    cell.font = NORMAL_FONT
    cell.border = THIN_BORDER
    cell.alignment = LEFT_ALIGN
    if is_alt:
        cell.fill = ALT_ROW_FILL
    if access_value == "read":
        cell.fill = READ_FILL
    elif access_value == "none":
        cell.fill = NONE_FILL

def set_column_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def create_workbook():
    wb = Workbook()

    # ========== Sheet 1: Metadata & Legend ==========
    ws_meta = wb.active
    ws_meta.title = "Metadata & Legend"

    ws_meta['A1'] = "Complete Page Views Inventory"
    ws_meta['A1'].font = TITLE_FONT

    ws_meta['A3'] = "Source of truth:"
    ws_meta['B3'] = "src/policy/security/identity/pageRegistry.ts (COMPONENT_GROUPS + PAGE_REGISTRY)"
    ws_meta['B3'].font = BOLD_FONT

    ws_meta['A4'] = "Generated from code inspection:"
    ws_meta['B4'] = "2026-06-10"

    ws_meta['A5'] = "Exported to spreadsheet:"
    ws_meta['B5'] = datetime.now().strftime("%Y-%m-%d")

    ws_meta['A7'] = "Tags Legend"
    ws_meta['A7'].font = BOLD_FONT

    legend = [
        ["[PAGE_REGISTRY]", "Controllable via the Page View Access matrix (admin tool for Marites/Robert) at /admin/users"],
        ["[FEATURE]", "Gated via FeatureRouteGuard / feature catalog (src/policy/security/features/catalog.ts)"],
        ["[ROUTE]", "Defined as a React Router <Route> in src/App.tsx"],
        ["[NAV]", "Appears in the main sidebar / mobile navigation (CommandCenterLayout)"],
    ]
    for i, (tag, desc) in enumerate(legend, 8):
        ws_meta[f'A{i}'] = tag
        ws_meta[f'A{i}'].font = BOLD_FONT
        ws_meta[f'B{i}'] = desc

    ws_meta['A13'] = "Default Access Color Key"
    ws_meta['A13'].font = BOLD_FONT
    ws_meta['A14'] = "read"
    ws_meta['A14'].fill = READ_FILL
    ws_meta['B14'] = "User can view by default (most pages)"
    ws_meta['A15'] = "none"
    ws_meta['A15'].fill = NONE_FILL
    ws_meta['B15'] = "User Management / admin pages — must be explicitly granted"

    ws_meta['A17'] = "Notes"
    ws_meta['A17'].font = BOLD_FONT
    notes = [
        "Most detail/sub-routes (e.g. /library/:policyId, /forms/:formId) inherit access from their parent page.",
        "User Management pages (cmp-user-management) default to 'none' (must be explicitly granted).",
        "The real matrix UI and guards live in PageAccessMatrix.tsx, pageAccess.ts, PageAccessRouteGuard.tsx.",
        "Re-generate this file after changes to pageRegistry.ts or App.tsx routes.",
    ]
    for i, note in enumerate(notes, 18):
        ws_meta[f'A{i}'] = note
        ws_meta.merge_cells(f'A{i}:F{i}')

    set_column_widths(ws_meta, [22, 90, 20, 20, 20, 20])

    # ========== Sheet 2: Component Groups ==========
    ws_groups = wb.create_sheet("Component Groups")

    ws_groups['A1'] = "Component Groups — The 11 modules in the Page View Access matrix"
    ws_groups['A1'].font = TITLE_FONT
    ws_groups.merge_cells('A1:D1')

    headers = ["componentId", "Label", "Default Access", "Order"]
    for col, h in enumerate(headers, 1):
        cell = ws_groups.cell(row=3, column=col, value=h)
        apply_header_style(cell)

    for row_idx, row in enumerate(COMPONENT_GROUPS, 4):
        is_alt = (row_idx % 2 == 0)
        for col_idx, val in enumerate(row, 1):
            cell = ws_groups.cell(row=row_idx, column=col_idx, value=val)
            apply_cell_style(cell, is_alt, access_value=row[2] if col_idx == 3 else None)
            if col_idx == 4:
                cell.alignment = CENTER_ALIGN

    set_column_widths(ws_groups, [24, 38, 16, 10])
    ws_groups.auto_filter.ref = f"A3:D{3 + len(COMPONENT_GROUPS)}"
    ws_groups.freeze_panes = "A4"

    # ========== Sheet 3: Page Views Master ==========
    ws_master = wb.create_sheet("Page Views Master")

    ws_master['A1'] = "Page Views Master Table — [PAGE_REGISTRY] entries (controllable in Page View Access matrix)"
    ws_master['A1'].font = TITLE_FONT
    ws_master.merge_cells('A1:H1')

    ws_master['A2'] = "Source: docs/Page_Views_List.md (2026-06-10) • All rows tagged [PAGE_REGISTRY]"
    ws_master['A2'].font = SUBTITLE_FONT
    ws_master.merge_cells('A2:H2')

    master_headers = ["pageId", "Label", "Route Pattern", "Component Group", "Default Access", "Fallback Feature", "Tags", "Nav Presence"]
    for col, h in enumerate(master_headers, 1):
        cell = ws_master.cell(row=4, column=col, value=h)
        apply_header_style(cell)

    for row_idx, row in enumerate(PAGE_VIEWS_MASTER, 5):
        is_alt = (row_idx % 2 == 0)
        for col_idx, val in enumerate(row, 1):
            cell = ws_master.cell(row=row_idx, column=col_idx, value=val)
            access_val = row[4]  # Default Access column
            apply_cell_style(cell, is_alt, access_value=access_val if col_idx == 5 else None)
            if col_idx in (1, 5):
                cell.alignment = CENTER_ALIGN

    set_column_widths(ws_master, [24, 28, 34, 22, 16, 28, 42, 22])
    ws_master.auto_filter.ref = f"A4:H{4 + len(PAGE_VIEWS_MASTER)}"
    ws_master.freeze_panes = "A5"
    ws_master.row_dimensions[1].height = 24

    # ========== Sheet 4: Additional Surfaces ==========
    ws_add = wb.create_sheet("Additional Surfaces")

    ws_add['A1'] = "Additional Surfaces — Not individually registered in [PAGE_REGISTRY] (inherit from parents or special)"
    ws_add['A1'].font = TITLE_FONT
    ws_add.merge_cells('A1:G1')

    add_headers = ["Key / Identifier", "Label", "Route Pattern", "Feature ID", "Tags", "Nav Presence", "Notes"]
    for col, h in enumerate(add_headers, 1):
        cell = ws_add.cell(row=3, column=col, value=h)
        apply_header_style(cell)

    for row_idx, row in enumerate(ADDITIONAL_SURFACES, 4):
        is_alt = (row_idx % 2 == 0)
        for col_idx, val in enumerate(row, 1):
            cell = ws_add.cell(row=row_idx, column=col_idx, value=val)
            apply_cell_style(cell, is_alt)
            if col_idx in (1, 4):
                cell.alignment = CENTER_ALIGN

    set_column_widths(ws_add, [28, 36, 42, 28, 22, 24, 24])
    ws_add.auto_filter.ref = f"A3:G{3 + len(ADDITIONAL_SURFACES)}"
    ws_add.freeze_panes = "A4"

    # ========== Sheet 5: All Combined (flat list for easy filtering) ==========
    ws_all = wb.create_sheet("All Page Views (Flat)")

    ws_all['A1'] = "ALL Page Views — Combined flat list (Master + Additional)"
    ws_all['A1'].font = TITLE_FONT
    ws_all.merge_cells('A1:I1')

    ws_all['A2'] = "Use filters to show only registry pages or additional surfaces. In Registry = TRUE for controllable pages."
    ws_all['A2'].font = SUBTITLE_FONT
    ws_all.merge_cells('A2:I2')

    all_headers = ["pageId / Key", "Label", "Route Pattern", "Component Group / Category", "Default Access", "Fallback Feature", "Tags", "Nav Presence", "In Page Registry"]
    for col, h in enumerate(all_headers, 1):
        cell = ws_all.cell(row=4, column=col, value=h)
        apply_header_style(cell)

    row_num = 5
    # Add master
    for row in PAGE_VIEWS_MASTER:
        vals = row + ["Yes"]
        is_alt = (row_num % 2 == 0)
        for col_idx, val in enumerate(vals, 1):
            cell = ws_all.cell(row=row_num, column=col_idx, value=val)
            apply_cell_style(cell, is_alt, access_value=vals[4] if col_idx == 5 else None)
        row_num += 1

    # Add additional
    for row in ADDITIONAL_SURFACES:
        vals = [row[0], row[1], row[2], "— (inherits)", "—", row[3], row[4], row[5], "No"]
        is_alt = (row_num % 2 == 0)
        for col_idx, val in enumerate(vals, 1):
            cell = ws_all.cell(row=row_num, column=col_idx, value=val)
            apply_cell_style(cell, is_alt)
        row_num += 1

    set_column_widths(ws_all, [28, 36, 42, 24, 16, 28, 42, 22, 16])
    ws_all.auto_filter.ref = f"A4:I{row_num - 1}"
    ws_all.freeze_panes = "A5"

    # Reorder sheets nicely
    wb._sheets = [ws_meta, ws_groups, ws_master, ws_add, ws_all]

    wb.save(OUTPUT)
    print(f"Created: {OUTPUT}")
    print(f"Sheets: {wb.sheetnames}")
    print(f"Total page views (master): {len(PAGE_VIEWS_MASTER)}")
    print(f"Total additional surfaces: {len(ADDITIONAL_SURFACES)}")
    print(f"Total rows in flat view: {len(PAGE_VIEWS_MASTER) + len(ADDITIONAL_SURFACES)}")

if __name__ == "__main__":
    create_workbook()
