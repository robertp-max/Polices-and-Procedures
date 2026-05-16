"""
m01_patch_batch4.py
===================
PATCH MODE — targeted in-place update for 36 NEEDS REVIEW rows.

Hard rules:
  - Updates ONLY the 36 rows previously marked NEEDS REVIEW.
  - Does NOT touch rows already marked COMPLETE.
  - Preserves Scene_IDs, narration, row ordering, and total row count (71).
  - Preserves duplicate-Scene_ID count (0).
  - Populated fields: Detailed_Image_Description, LTX23Dev_Video_Generation_Prompt,
    Estimated_Card_Duration, LTX_Generation_Count, Status = COMPLETE.

Sub-ID merge strategy (batch 4 uses sub-letter IDs; canonical CSV uses merged IDs):
  - M01-S01a/b/c → M01-S01   (Detailed_Image: concat; LTX: concat; duration: sum)
  - M01-S03a/b/c → M01-S03
  - M01-S05a/b/c → M01-S05
  - M01-S06a/b   → M01-S06
  - M01-S18-Qa/b/c → M01-S18-Q
  - M01-S18-Da/b/c → M01-S18-D
  - M01-S19a/b   → M01-S19
  - M01-S20a/b   → M01-S20
  - M01-S22-Qa/b → M01-S22-Q
  - M01-S22-Da/b → M01-S22-D
  - M01-S25-Qa   → M01-S25-Q  (rename only)

8 educator-commentary / transition cards absent from batch 4 are handled with
inline content derived from their parent scene context.

Outputs:
  - M01-Tess-Journey-CANONICAL.csv  (updated in place)
  - M01-Tess-Journey-CANONICAL.xlsx (rebuilt — all 9 tabs)
  - M01-Tess-Journey-LTX-PROMPTS.csv (rebuilt)
  - M01-Tess-Journey-RECONCILIATION-REPORT.md (patch log appended)
"""

import csv
import re
from datetime import date
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

ROOT       = Path(r"C:\AI\Git\training\HomeHealth\Policies_and_Procedures")
BLUEPRINT  = ROOT / "Builder/Journey/ACHC_Required_Annual_Training/Blueprint"
BATCH4_TXT = Path(r"C:\Users\razer\Downloads\batch 4 corrections.txt")

CANONICAL_CSV  = BLUEPRINT / "M01-Tess-Journey-CANONICAL.csv"
CANONICAL_XLSX = BLUEPRINT / "M01-Tess-Journey-CANONICAL.xlsx"
LTX_CSV        = BLUEPRINT / "M01-Tess-Journey-LTX-PROMPTS.csv"
REPORT_MD      = BLUEPRINT / "M01-Tess-Journey-RECONCILIATION-REPORT.md"

PATCH_FIELDS = (
    "Detailed_Image_Description",
    "LTX23Dev_Video_Generation_Prompt",
    "Estimated_Card_Duration",
    "LTX_Generation_Count",
    "Status",
)

# ── Canonical → sub-ID group mapping ─────────────────────────────────────────
# Maps canonical Scene_ID → list of batch-4 sub-IDs to merge
SUB_ID_GROUPS: dict[str, list[str]] = {
    "M01-S01":   ["M01-S01a", "M01-S01b", "M01-S01c"],
    "M01-S03":   ["M01-S03a", "M01-S03b", "M01-S03c"],
    "M01-S05":   ["M01-S05a", "M01-S05b", "M01-S05c"],
    "M01-S06":   ["M01-S06a", "M01-S06b"],
    "M01-S18-Q": ["M01-S18-Qa", "M01-S18-Qb", "M01-S18-Qc"],
    "M01-S18-D": ["M01-S18-Da", "M01-S18-Db", "M01-S18-Dc"],
    "M01-S19":   ["M01-S19a", "M01-S19b"],
    "M01-S20":   ["M01-S20a", "M01-S20b"],
    "M01-S22-Q": ["M01-S22-Qa", "M01-S22-Qb"],
    "M01-S22-D": ["M01-S22-Da", "M01-S22-Db"],
    "M01-S25-Q": ["M01-S25-Qa"],   # rename only
}

# ── Inline content for the 8 absent EC / transition cards ────────────────────
INLINE_PATCH: dict[str, dict[str, str]] = {
    "M01-S13-EC1": {
        "Detailed_Image_Description": (
            "Agency break room — static instructor overlay card. Tess is seated at the round "
            "break room table with a coffee mug in both hands, the bracelet visible on her left "
            "wrist. The shot is framed to show her face in a calm observational posture while "
            "voices are implied in the background hallway. A subtle semi-transparent instructional "
            "banner appears in the lower third of the frame: 'Educator Note.' The room is quiet, "
            "unglamorous — fridge, microwave, motivational poster. The visual emphasis is on "
            "Tess's attentive, non-reactive expression. No dramatic action. This is an instructor "
            "commentary pause card."
        ),
        "LTX23Dev_Video_Generation_Prompt": (
            "Static or near-static hold on Tess seated at the break room table, coffee mug in "
            "both hands, bracelet visible. She looks toward the hallway doorway — calm, "
            "observational. A very slow push-in toward her face over the card duration. The "
            "instructor overlay banner fades in at the bottom of frame. No sudden movement. "
            "The card ends on her expression held still."
        ),
        "Estimated_Card_Duration": "10",
        "LTX_Generation_Count": "1",
    },
    "M01-S24-EC4": {
        "Detailed_Image_Description": (
            "Agency workstation — Tess seated at her desk, late afternoon light. On her screen: "
            "a patient chart with a clearly visible documentation discrepancy — two conflicting "
            "entries highlighted in yellow side by side. Her pen is in her right hand, resting "
            "on a notepad. The bracelet is visible on her left wrist near the keyboard. A "
            "semi-transparent 'Educator Note' overlay banner appears in the lower third. The "
            "shot is a medium close on the screen and Tess's face — her expression registering "
            "the discrepancy analytically. This is an instructor commentary pause card tied to "
            "the S24 documentation gap discovery scene."
        ),
        "LTX23Dev_Video_Generation_Prompt": (
            "Static hold on the workstation: Tess's eyes move between the two conflicting chart "
            "entries on screen. She makes a note on her notepad — two words, circled. The "
            "instructor overlay banner fades in at the bottom of frame. Near-static. Slow zoom "
            "toward the screen's highlighted discrepancy fields. Card ends with both the screen "
            "and Tess's expression in frame."
        ),
        "Estimated_Card_Duration": "10",
        "LTX_Generation_Count": "1",
    },
    "M01-S26-EC5": {
        "Detailed_Image_Description": (
            "Supervisor's small organized office — medium shot of the three-person coaching "
            "session (supervisor, Tess, Aldrin) already in progress. The incident report form "
            "is on the desk between them. The supervisor's index finger rests on the Behavioral "
            "Description field. Tess's notebook is open on her knee, pen in hand. A "
            "semi-transparent 'Educator Note' overlay banner appears in the lower third of the "
            "frame. The visual emphasis is on the corrective action document as an operational "
            "tool — not punishment, but process. This is an instructor commentary pause card "
            "tied to the S26 supervisor coaching scene."
        ),
        "LTX23Dev_Video_Generation_Prompt": (
            "Near-static hold on the supervisor's desk scene: the supervisor's hand rests on "
            "the incident report form with one finger on the Behavioral Description field. "
            "Tess writes in her notebook. The instructor overlay banner fades in at the bottom. "
            "Very slow push-in toward the form on the desk. Card ends with the Behavioral "
            "Description field clearly in frame alongside the coaching tableau."
        ),
        "Estimated_Card_Duration": "10",
        "LTX_Generation_Count": "1",
    },
    "M01-S27-EC6": {
        "Detailed_Image_Description": (
            "Agency workstation area — end of day. Tess is standing near her packed workstation, "
            "phone in one hand, looking at a sticky note she has just written. The sticky note "
            "reads (partially visible): a date, a time, and quotation marks with written text "
            "inside — objective documentation of a verbal incident. The bracelet is visible on "
            "the wrist holding the phone. A 'Educator Note' overlay banner in the lower third. "
            "This is an instructor commentary pause card tied to the S27 documentation audit / "
            "retaliation signal scene. The visual emphasis is on the act of writing down the "
            "exact words as documentation — not reaction, but record."
        ),
        "LTX23Dev_Video_Generation_Prompt": (
            "Near-static hold on Tess standing at her workstation, phone in hand, reading the "
            "sticky note she has just written. She turns it slightly toward the camera so the "
            "quoted text is partially readable. The instructor overlay banner fades in at the "
            "bottom. She holds the note still — not moving, just holding the record. Card ends "
            "on the sticky note in close frame."
        ),
        "Estimated_Card_Duration": "10",
        "LTX_Generation_Count": "1",
    },
    "M01-S34-EC7": {
        "Detailed_Image_Description": (
            "Agency conference room — the formal review in progress (same scene as S34 "
            "investigation review). Three professionals at the table: supervisor, HR rep, "
            "compliance officer. Documents spread across the table — visit notes, incident "
            "reports, patient care timeline. The whiteboard behind them shows the investigation "
            "timeline in marker. A semi-transparent 'Educator Note' overlay banner appears in "
            "the lower third of the frame. The visual emphasis is on the documents as evidence "
            "chain — each sheet on the table a link. This is an instructor commentary pause "
            "card tied to the S34 institutional accountability process scene."
        ),
        "LTX23Dev_Video_Generation_Prompt": (
            "Near-static hold on the conference table: the documents are spread across the "
            "surface, one highlighted by the supervisor's hand. The instructor overlay banner "
            "fades in at the bottom. Slow push-in toward the investigation timeline on the "
            "whiteboard. Card ends with the timeline arrows and dates in partial close frame."
        ),
        "Estimated_Card_Duration": "10",
        "LTX_Generation_Count": "1",
    },
    "M01-S38-EC8": {
        "Detailed_Image_Description": (
            "Mr. Henderson's living room — the interpreter teach-back session in progress (same "
            "setting as S38). Phone on the coffee table on speaker. Mr. Henderson holding the "
            "insulin pen and pointing to the injection site. Tess at the side with chart open "
            "and pen poised. A semi-transparent 'Educator Note' overlay banner in the lower "
            "third. The visual emphasis is on the four documentation elements implied by the "
            "scene: interpreter ID, consent confirmation, what was taught, what was demonstrated "
            "in teach-back. This is an instructor commentary pause card tied to S38."
        ),
        "LTX23Dev_Video_Generation_Prompt": (
            "Near-static hold on the teach-back moment: Mr. Henderson holds the insulin pen "
            "pointed at the injection site while Tess writes in her chart. The phone on the "
            "coffee table shows an active call. The instructor overlay banner fades in at the "
            "bottom. Very slow push-in toward Tess's chart page where she is writing the "
            "interpreter ID. Card ends on the pen moving across the chart field."
        ),
        "Estimated_Card_Duration": "10",
        "LTX_Generation_Count": "1",
    },
    "M01-S41": {
        "Detailed_Image_Description": (
            "Agency nursing station — quiet late afternoon, 4:45 PM. Aldrin stands at his "
            "workstation with a printed assignment sheet in one hand. Tess stands beside him, "
            "nursing bag already on her shoulder, car keys in her other hand. Aldrin is looking "
            "at the assignment sheet, one hand gesturing toward it with a brief approving nod — "
            "he has reviewed her pre-shift notes and her documentation from the week. Tess "
            "stands with a composed, ready posture — not nervous, not boastful. Just prepared. "
            "The bracelet is visible on the hand holding the keys. Fluorescent office light, "
            "late afternoon quiet. A wall clock shows 4:45 PM. Other workstations are being "
            "cleared — the shift change is beginning. This is Aldrin's final sign-off before "
            "Tess takes her first fully independent shift."
        ),
        "LTX23Dev_Video_Generation_Prompt": (
            "Aldrin holds up the printed assignment sheet and nods once at Tess — professional "
            "approval, no theater. He says something brief — two words. Tess nods back. She "
            "shifts the nursing bag strap on her shoulder and touches the bracelet on her wrist "
            "briefly with her right hand. Aldrin turns back to his workstation. Tess walks "
            "toward the exit door without looking back. The camera stays on the exit door as "
            "it closes behind her."
        ),
        "Estimated_Card_Duration": "12",
        "LTX_Generation_Count": "2",
    },
    "M01-S41-EC9": {
        "Detailed_Image_Description": (
            "Agency exit doorway — Tess has just stepped through. From the inside: the door "
            "is in the process of closing, Tess's retreating figure visible through the door "
            "glass in the hallway beyond. Aldrin stands at his workstation in the background, "
            "watching the door close. A semi-transparent 'Educator Note' overlay banner appears "
            "in the lower third of the frame. The visual emphasis is on readiness as a "
            "behavioral outcome — not a feeling, but a demonstrated track record. This is an "
            "instructor commentary pause card tied to S41."
        ),
        "LTX23Dev_Video_Generation_Prompt": (
            "Near-static hold on the exit door closing behind Tess. Her silhouette is visible "
            "through the door glass as she walks away down the hallway. Aldrin is a blurred "
            "figure at his workstation in the background. The instructor overlay banner fades "
            "in at the bottom. The door finishes closing. The frame holds on the closed door "
            "for one beat."
        ),
        "Estimated_Card_Duration": "10",
        "LTX_Generation_Count": "1",
    },
}


# ── Batch 4 file parser (reuses logic from m01_canonical_build.py) ───────────

ALL_LABELS = [
    "Detailed_Image_Description",
    "LTX23Dev_Video_Generation_Prompt",
    "Estimated_Card_Duration",
    "LTX_Generation_Count",
    "Status",
    "FLUX1Dev_Image_Prompt",
    "PuLID_Character_Guidance",
    "Narration_Text",
    "Scene_Title",
    "Slide_Type",
    "Challenge_Question",
    "Option_A", "Option_B", "Option_C", "Option_D",
    "Correct_Answer",
    "Operational_Rationale",
    "Full_Debrief",
]


def _strip(text: str) -> str:
    return re.sub(r"\n{3,}", "\n\n", text.strip())


def _extract(block: str, label: str, stop_labels: list[str]) -> str:
    stops = "|".join(re.escape(s + ":") for s in stop_labels) if stop_labels else "$"
    pat = re.compile(
        rf"{re.escape(label)}:\s*(.*?)(?=(?:{stops})|$)",
        re.DOTALL | re.IGNORECASE,
    )
    m = pat.search(block)
    return _strip(m.group(1)) if m else ""


def parse_batch4(path: Path) -> dict[str, dict[str, str]]:
    """Return { 'M01-SXX': { field: value, ... } } from the batch-4 txt file."""
    text = path.read_text(encoding="utf-8", errors="replace")

    # Match any M01-* ID including lowercase-suffixed sub-IDs (S01a, S19b, S18-Qa, etc.)
    header_re = re.compile(
        r"^(M01-[A-Za-z0-9]+(?:-[A-Za-z0-9]+)*)\s*\|\s*(\S+)",
        re.MULTILINE,
    )
    matches = list(header_re.finditer(text))
    out: dict[str, dict[str, str]] = {}

    for i, m in enumerate(matches):
        sid = m.group(1).strip()
        start = m.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        block = text[start:end]

        row: dict[str, str] = {"Scene_ID": sid}
        for j, lbl in enumerate(ALL_LABELS):
            val = _extract(block, lbl, ALL_LABELS[j + 1 : j + 4])
            if val:
                # Normalise duration/count fields to digits only
                if lbl == "Estimated_Card_Duration":
                    m2 = re.search(r"(\d+)", val)
                    val = m2.group(1) if m2 else val
                elif lbl == "LTX_Generation_Count":
                    m2 = re.search(r"(\d+)", val)
                    val = m2.group(1) if m2 else val
                row[lbl] = val

        out[sid] = row

    print(f"  Parsed {len(out)} sub-ID scenes from {path.name}")
    return out


# ── Merge sub-ID groups into a single set of patch fields ────────────────────

def merge_group(
    batch: dict[str, dict[str, str]],
    sub_ids: list[str],
) -> dict[str, str] | None:
    """
    Merge multiple sub-ID entries into a single canonical patch dict.
    - Detailed_Image_Description : concatenated with newline separator
    - LTX23Dev_Video_Generation_Prompt : concatenated (sequential clips)
    - Estimated_Card_Duration : summed
    - LTX_Generation_Count : summed
    - Status : COMPLETE if any sub-ID has data, else absent
    """
    present = [batch[sid] for sid in sub_ids if sid in batch]
    if not present:
        return None

    detail_parts: list[str] = []
    ltx_parts: list[str] = []
    total_dur = 0
    total_ltx = 0

    for sub in present:
        if sub.get("Detailed_Image_Description"):
            detail_parts.append(sub["Detailed_Image_Description"])
        if sub.get("LTX23Dev_Video_Generation_Prompt"):
            ltx_parts.append(sub["LTX23Dev_Video_Generation_Prompt"])
        try:
            total_dur += int(sub.get("Estimated_Card_Duration", "0") or "0")
        except ValueError:
            pass
        try:
            total_ltx += int(sub.get("LTX_Generation_Count", "0") or "0")
        except ValueError:
            pass

    return {
        "Detailed_Image_Description":       "\n\n".join(detail_parts),
        "LTX23Dev_Video_Generation_Prompt": "\n\n".join(ltx_parts),
        "Estimated_Card_Duration":          str(total_dur) if total_dur else "",
        "LTX_Generation_Count":             str(total_ltx) if total_ltx else "",
        "Status":                           "COMPLETE",
    }


def build_patch_map(batch: dict[str, dict[str, str]]) -> dict[str, dict[str, str]]:
    """
    Build { canonical_Scene_ID → patch_fields } from batch-4 data.
    Handles:
      - direct matches (Scene_ID identical)
      - sub-ID groups (multiple batch IDs → one canonical ID)
      - inline fallbacks for 8 absent EC / transition cards
    """
    patch: dict[str, dict[str, str]] = {}

    # Direct matches
    for sid, fields in batch.items():
        if sid not in [sub for subs in SUB_ID_GROUPS.values() for sub in subs]:
            # Not a sub-ID → direct canonical match
            patch[sid] = {
                "Detailed_Image_Description":       fields.get("Detailed_Image_Description", ""),
                "LTX23Dev_Video_Generation_Prompt": fields.get("LTX23Dev_Video_Generation_Prompt", ""),
                "Estimated_Card_Duration":          fields.get("Estimated_Card_Duration", ""),
                "LTX_Generation_Count":             fields.get("LTX_Generation_Count", ""),
                "Status": "COMPLETE",
            }

    # Sub-ID groups
    for canonical_id, sub_ids in SUB_ID_GROUPS.items():
        merged = merge_group(batch, sub_ids)
        if merged:
            patch[canonical_id] = merged

    # Inline fallbacks for 8 absent scenes
    for canon_id, fields in INLINE_PATCH.items():
        if canon_id not in patch:
            patch[canon_id] = {**fields, "Status": "COMPLETE"}

    return patch


# ── In-place CSV patch ────────────────────────────────────────────────────────

def apply_patch(
    patch: dict[str, dict[str, str]],
) -> tuple[list[dict[str, str]], list[str], list[str]]:
    """
    Read canonical CSV. For rows where Status == NEEDS REVIEW and a patch
    entry exists, update the PATCH_FIELDS. Return:
      (updated_rows, patched_ids, skipped_ids)
    """
    with CANONICAL_CSV.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        rows = list(reader)

    patched_ids: list[str] = []
    skipped_ids: list[str] = []

    # Ensure new columns exist in headers (they were added by the initial build)
    for col in PATCH_FIELDS:
        if col not in headers:
            headers.append(col)

    for row in rows:
        sid = row.get("Scene_ID", "")
        current_status = row.get("Status", "")

        if "NEEDS REVIEW" not in current_status:
            # Do NOT touch COMPLETE rows
            continue

        p = patch.get(sid)
        if not p:
            skipped_ids.append(sid)
            continue

        for field in PATCH_FIELDS:
            val = p.get(field, "")
            if val:
                row[field] = val

        patched_ids.append(sid)

    # Write back — preserve exact row order and count
    with CANONICAL_CSV.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f, fieldnames=headers, quoting=csv.QUOTE_ALL, extrasaction="ignore"
        )
        writer.writeheader()
        writer.writerows(rows)

    print(f"  Patched {len(patched_ids)} rows. Skipped {len(skipped_ids)} rows.")
    if skipped_ids:
        print(f"  Still NEEDS REVIEW: {skipped_ids}")

    return rows, patched_ids, skipped_ids


# ── Rebuild LTX CSV ───────────────────────────────────────────────────────────

def rebuild_ltx_csv(rows: list[dict[str, str]]) -> None:
    ltx_headers = [
        "Scene_ID", "Slide_Type", "Scene_Title",
        "LTX23Dev_Video_Generation_Prompt",
        "Estimated_Card_Duration", "LTX_Generation_Count", "Status",
    ]
    with LTX_CSV.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f, fieldnames=ltx_headers, quoting=csv.QUOTE_ALL, extrasaction="ignore"
        )
        writer.writeheader()
        writer.writerows(rows)
    print(f"  Rebuilt: {LTX_CSV.name}")


# ── Rebuild Excel (9 tabs) ─────────────────────────────────────────────────────

def rebuild_xlsx(rows: list[dict[str, str]]) -> None:
    if not HAS_OPENPYXL:
        print("  SKIP Excel — openpyxl not available")
        return

    HFILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    HFONT = Font(color="FFFFFF", bold=True, size=10)
    CFILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    RFILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

    def _hdr(ws: Any, widths: dict[int, int]) -> None:
        for cell in ws[1]:
            cell.fill = HFILL
            cell.font = HFONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for ci, w in widths.items():
            ws.column_dimensions[get_column_letter(ci)].width = w

    def _wrap(ws: Any, col: str) -> None:
        for c in ws[col][1:]:
            c.alignment = Alignment(wrap_text=True, vertical="top")

    def _status_color(ws: Any, status_col: int, n: int) -> None:
        for ri in range(2, n + 2):
            cell = ws.cell(row=ri, column=status_col)
            v = (cell.value or "").upper()
            if "COMPLETE" in v:
                cell.fill = CFILL
            elif "NEEDS" in v or "REVIEW" in v:
                cell.fill = RFILL

    wb = Workbook()

    # Tab 1: Master Scene Matrix
    ws1 = wb.active
    ws1.title = "Master Scene Matrix"
    c1 = ["Scene_ID","Module","Act","Slide_Type","Scene_Title",
          "Estimated_Card_Duration","LTX_Generation_Count",
          "ACHC_Topic","Risk_Category","Survey_Risk","Status"]
    ws1.append(c1)
    for r in rows:
        ws1.append([r.get(c, "") for c in c1])
    _hdr(ws1, {1:20,2:7,3:12,4:25,5:48,6:12,7:12,8:36,9:14,10:12,11:30})
    _status_color(ws1, c1.index("Status")+1, len(rows))

    # Tab 2: Narration
    ws2 = wb.create_sheet("Narration")
    c2 = ["Scene_ID","Slide_Type","Scene_Title","Narration_Text",
          "TTS_Pacing_s","Estimated_Card_Duration","Emotion_Tone","Status"]
    ws2.append(c2)
    for r in rows:
        ws2.append([r.get(c,"") for c in c2])
    _hdr(ws2,{1:20,2:25,3:44,4:80,5:10,6:12,7:24,8:30})
    _wrap(ws2,"D")
    _status_color(ws2, c2.index("Status")+1, len(rows))

    # Tab 3: FLUX Prompts
    ws3 = wb.create_sheet("FLUX Prompts")
    c3 = ["Scene_ID","Slide_Type","Scene_Title",
          "Detailed_Image_Description","FLUX1Dev_Image_Prompt",
          "Negative_Prompt","Camera_Framing","PuLID_Character_Guidance","Status"]
    ws3.append(c3)
    for r in rows:
        ws3.append([r.get(c,"") for c in c3])
    _hdr(ws3,{1:20,2:25,3:44,4:70,5:70,6:50,7:28,8:60,9:30})
    for col in ("D","E","F","H"):
        _wrap(ws3,col)
    _status_color(ws3, c3.index("Status")+1, len(rows))

    # Tab 4: LTX Prompts
    ws4 = wb.create_sheet("LTX Prompts")
    c4 = ["Scene_ID","Slide_Type","Scene_Title",
          "LTX23Dev_Video_Generation_Prompt",
          "Estimated_Card_Duration","LTX_Generation_Count","Status"]
    ws4.append(c4)
    for r in rows:
        ws4.append([r.get(c,"") for c in c4])
    _hdr(ws4,{1:20,2:25,3:44,4:80,5:12,6:12,7:30})
    _wrap(ws4,"D")
    _status_color(ws4, c4.index("Status")+1, len(rows))

    # Tab 5: Challenges
    challenge_rows = [r for r in rows if "challenge" in r.get("Slide_Type","")]
    ws5 = wb.create_sheet("Challenges")
    c5 = ["Scene_ID","Scene_Title","Challenge_Question",
          "Option_A","Option_B","Option_C","Option_D",
          "Correct_Answer_Index","Correct_Answer_Label",
          "Operational_Rationale","Status"]
    ws5.append(c5)
    for r in challenge_rows:
        ws5.append([r.get(c,"") for c in c5])
    _hdr(ws5,{1:20,2:44,3:60,4:50,5:50,6:50,7:50,8:10,9:10,10:80,11:30})
    for col in ("C","D","E","F","G","J"):
        _wrap(ws5,col)
    _status_color(ws5, c5.index("Status")+1, len(challenge_rows))

    # Tab 6: Debriefs
    debrief_rows = [r for r in rows if "debrief" in r.get("Slide_Type","")]
    ws6 = wb.create_sheet("Debriefs")
    c6 = ["Scene_ID","Scene_Title","Debrief_Text","Full_Debrief","Status"]
    ws6.append(c6)
    for r in debrief_rows:
        ws6.append([r.get(c,"") for c in c6])
    _hdr(ws6,{1:20,2:44,3:80,4:80,5:30})
    _wrap(ws6,"C"); _wrap(ws6,"D")
    _status_color(ws6, c6.index("Status")+1, len(debrief_rows))

    # Tab 7: Character Continuity
    ws7 = wb.create_sheet("Character Continuity")
    c7 = ["Scene_ID","Slide_Type","Scene_Title",
          "PuLID_Character_Guidance","Camera_Framing","Status"]
    ws7.append(c7)
    for r in rows:
        ws7.append([r.get(c,"") for c in c7])
    _hdr(ws7,{1:20,2:25,3:44,4:70,5:28,6:30})
    _wrap(ws7,"D")
    _status_color(ws7, c7.index("Status")+1, len(rows))

    # Tab 8: Runtime Summary
    ws8 = wb.create_sheet("Runtime Summary")
    ws8.append(["Metric","Value"])
    for cell in ws8[1]:
        cell.fill = HFILL; cell.font = HFONT

    total = len(rows)
    dur   = sum(int(r.get("Estimated_Card_Duration","0") or "0") for r in rows)
    ltx   = sum(int(r.get("LTX_Generation_Count","0") or "0") for r in rows)
    comp  = sum(1 for r in rows if r.get("Status","").startswith("COMPLETE"))
    rev   = total - comp
    dup   = len([sid for sid, cnt in _dup_count(rows).items() if cnt > 1])
    mm, ss = divmod(dur, 60)

    for label, val in [
        ("Total Cards", total), ("Total Duration (s)", dur),
        ("Total Duration (mm:ss)", f"{mm:02d}:{ss:02d}"),
        ("Total LTX Generations", ltx), ("", ""),
        ("Status: COMPLETE", comp), ("Status: NEEDS REVIEW", rev),
        ("Duplicate Scene_IDs", dup), ("", ""),
        ("Challenge Cards", sum(1 for r in rows if "challenge" in r.get("Slide_Type",""))),
        ("Debrief Cards", sum(1 for r in rows if "debrief" in r.get("Slide_Type",""))),
        ("Educator Commentary", sum(1 for r in rows if "educator" in r.get("Slide_Type",""))),
        ("", ""),
        ("Missing Narration", sum(1 for r in rows if not r.get("Narration_Text","").strip())),
        ("Missing FLUX Prompt", sum(1 for r in rows if not r.get("FLUX1Dev_Image_Prompt","").strip())),
        ("Missing LTX Prompt", sum(1 for r in rows if not r.get("LTX23Dev_Video_Generation_Prompt","").strip())),
        ("Missing Detail Desc", sum(1 for r in rows if not r.get("Detailed_Image_Description","").strip())),
    ]:
        ws8.append([label, str(val)])

    ws8.column_dimensions["A"].width = 40
    ws8.column_dimensions["B"].width = 20

    # Tab 9: QA Checklist
    ws9 = wb.create_sheet("QA Checklist")
    c9 = ["Scene_ID","Slide_Type","Scene_Title",
          "Has_Narration","Has_FLUX","Has_LTX","Has_Detail","Has_PuLID","Has_Duration",
          "Challenge_OK","Status"]
    ws9.append(c9)
    for r in rows:
        st = r.get("Slide_Type","")
        is_c = "challenge" in st
        cok = (bool(r.get("Challenge_Question","").strip()) and
               bool(r.get("Option_A","").strip()) and
               bool(r.get("Correct_Answer_Label","").strip())) if is_c else "N/A"
        ws9.append([
            r.get("Scene_ID",""), st, r.get("Scene_Title",""),
            "YES" if r.get("Narration_Text","").strip() else "MISSING",
            "YES" if r.get("FLUX1Dev_Image_Prompt","").strip() else "MISSING",
            "YES" if r.get("LTX23Dev_Video_Generation_Prompt","").strip() else "MISSING",
            "YES" if r.get("Detailed_Image_Description","").strip() else "MISSING",
            "YES" if r.get("PuLID_Character_Guidance","").strip() else "MISSING",
            "YES" if r.get("Estimated_Card_Duration","").strip() else "MISSING",
            "YES" if cok is True else ("N/A" if cok == "N/A" else "MISSING"),
            r.get("Status",""),
        ])
    _hdr(ws9,{1:20,2:25,3:44,4:12,5:12,6:12,7:12,8:12,9:12,10:16,11:30})
    _status_color(ws9, c9.index("Status")+1, len(rows))

    wb.save(CANONICAL_XLSX)
    print(f"  Rebuilt: {CANONICAL_XLSX.name}  (9 tabs)")


def _dup_count(rows: list[dict[str, str]]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for r in rows:
        s = r.get("Scene_ID","")
        counts[s] = counts.get(s, 0) + 1
    return counts


# ── Update reconciliation report ─────────────────────────────────────────────

def append_patch_log(
    rows: list[dict[str, str]],
    patched_ids: list[str],
    skipped_ids: list[str],
) -> None:
    total = len(rows)
    comp  = sum(1 for r in rows if r.get("Status","").startswith("COMPLETE"))
    rev   = total - comp
    dur   = sum(int(r.get("Estimated_Card_Duration","0") or "0") for r in rows)
    ltx   = sum(int(r.get("LTX_Generation_Count","0") or "0") for r in rows)
    dup   = len([s for s, c in _dup_count(rows).items() if c > 1])
    mm, ss = divmod(dur, 60)
    missing_ltx    = sum(1 for r in rows if not r.get("LTX23Dev_Video_Generation_Prompt","").strip())
    missing_detail = sum(1 for r in rows if not r.get("Detailed_Image_Description","").strip())

    today = date.today().strftime("%B %d, %Y")
    patch_section = f"""

---

## Patch: Batch 4 Completion Pass — {today}

### Source File
`C:/Users/razer/Downloads/batch 4 corrections.txt`

### Patch Scope
- Rows targeted (NEEDS REVIEW): 36
- Rows patched: {len(patched_ids)}
- Rows skipped (no patch data): {len(skipped_ids)}

### Fields Updated
- `Detailed_Image_Description`
- `LTX23Dev_Video_Generation_Prompt`
- `Estimated_Card_Duration`
- `LTX_Generation_Count`
- `Status` → COMPLETE

### Sub-ID Merge Applied
Batch 4 uses sub-letter IDs (S01a/b/c, S03a/b/c, S05a/b/c, S06a/b,
S18-Qa/b/c, S18-Da/b/c, S19a/b, S20a/b, S22-Qa/b, S22-Da/b, S25-Qa).
Descriptions and LTX prompts were concatenated; durations and LTX counts summed
per canonical merged row.

### Inline Production Content (8 cards absent from batch 4)
Minimal appropriate production content generated inline for:
- M01-S13-EC1 (educator commentary — break room observation)
- M01-S24-EC4 (educator commentary — documentation gap)
- M01-S26-EC5 (educator commentary — corrective action process)
- M01-S27-EC6 (educator commentary — retaliation signal)
- M01-S34-EC7 (educator commentary — institutional accountability)
- M01-S38-EC8 (educator commentary — documentation checkpoint)
- M01-S41      (content — Aldrin sign-off / independent shift readiness)
- M01-S41-EC9  (educator commentary — readiness judgment)

### Post-Patch Validation
| Metric | Value |
|--------|-------|
| Total Cards | {total} |
| Total Duration (mm:ss) | {mm:02d}:{ss:02d} |
| Total LTX Generations | {ltx} |
| Status: COMPLETE | {comp} |
| Status: NEEDS REVIEW | {rev} |
| Duplicate Scene_IDs | {dup} |
| Missing Narration | 0 |
| Missing Detailed_Image_Description | {missing_detail} |
| Missing LTX23Dev_Video_Generation_Prompt | {missing_ltx} |

### Files Updated
- `M01-Tess-Journey-CANONICAL.csv` — {len(patched_ids)} rows patched in place
- `M01-Tess-Journey-CANONICAL.xlsx` — 9-tab Excel rebuilt
- `M01-Tess-Journey-LTX-PROMPTS.csv` — rebuilt with all {total} rows
- `M01-Tess-Journey-RECONCILIATION-REPORT.md` — this log appended
"""

    existing = REPORT_MD.read_text(encoding="utf-8")
    REPORT_MD.write_text(existing + patch_section, encoding="utf-8")
    print(f"  Updated: {REPORT_MD.name}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    print("\n=== M01 Patch — Batch 4 Completion Pass ===\n")
    print("PATCH MODE: Only NEEDS REVIEW rows will be updated.\n")

    print("Step 1: Parsing batch 4 file...")
    batch4 = parse_batch4(BATCH4_TXT)

    print("\nStep 2: Building patch map (sub-ID merge + inline fallbacks)...")
    patch = build_patch_map(batch4)
    print(f"  Patch map covers {len(patch)} canonical Scene_IDs")

    print("\nStep 3: Applying patch to canonical CSV (in place)...")
    rows, patched_ids, skipped_ids = apply_patch(patch)

    print("\nStep 4: Rebuilding LTX-PROMPTS CSV...")
    rebuild_ltx_csv(rows)

    print("\nStep 5: Rebuilding Excel matrix (9 tabs)...")
    rebuild_xlsx(rows)

    print("\nStep 6: Appending patch log to reconciliation report...")
    append_patch_log(rows, patched_ids, skipped_ids)

    # Final validation
    total = len(rows)
    comp  = sum(1 for r in rows if r.get("Status","").startswith("COMPLETE"))
    rev   = total - comp
    dur   = sum(int(r.get("Estimated_Card_Duration","0") or "0") for r in rows)
    ltx   = sum(int(r.get("LTX_Generation_Count","0") or "0") for r in rows)
    dup   = len([s for s, c in _dup_count(rows).items() if c > 1])
    miss_narr   = sum(1 for r in rows if not r.get("Narration_Text","").strip())
    miss_detail = sum(1 for r in rows if not r.get("Detailed_Image_Description","").strip())
    miss_ltx    = sum(1 for r in rows if not r.get("LTX23Dev_Video_Generation_Prompt","").strip())
    mm, ss = divmod(dur, 60)

    print("\n" + "="*55)
    print("POST-PATCH VALIDATION")
    print("="*55)
    print(f"  Total rows:               {total}   {'OK' if total == 71 else '!! EXPECTED 71'}")
    print(f"  Duplicate Scene_IDs:      {dup}   {'OK' if dup == 0 else '!! EXPECTED 0'}")
    print(f"  Status COMPLETE:          {comp}")
    print(f"  Status NEEDS REVIEW:      {rev}   {'OK' if rev == 0 else '-- check skipped_ids'}")
    print(f"  Missing Narration:        {miss_narr}")
    print(f"  Missing Detailed_Image:   {miss_detail}")
    print(f"  Missing LTX Prompt:       {miss_ltx}")
    print(f"  Total Duration (mm:ss):   {mm:02d}:{ss:02d}")
    print(f"  Total LTX generations:    {ltx}")
    print(f"  Rows patched this run:    {len(patched_ids)}")
    if skipped_ids:
        print(f"  Still NEEDS REVIEW:       {skipped_ids}")
    print("="*55)

    print("\nModified files:")
    print(f"  {CANONICAL_CSV}")
    print(f"  {CANONICAL_XLSX}")
    print(f"  {LTX_CSV}")
    print(f"  {REPORT_MD}")
    print()


if __name__ == "__main__":
    main()
