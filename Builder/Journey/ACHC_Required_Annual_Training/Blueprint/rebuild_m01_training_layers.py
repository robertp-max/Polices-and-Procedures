import csv
import json
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(r"C:\AI\Git\training\HomeHealth\Policies_and_Procedures")
CSV_PATH = ROOT / "Builder/Journey/ACHC_Required_Annual_Training/Blueprint/M01-Marites-Journey-Narration-Defensibility-COMPLETE.csv"
XLSX_PATH = ROOT / "Builder/Journey/ACHC_Required_Annual_Training/Blueprint/M01-Marites-Journey-Narration-Defensibility-COMPLETE.xlsx"
TS_PATH = ROOT / "src/policy/journey/data/stagingM01Slides.ts"
PIPELINE_DIR = ROOT / "Builder/Journey/ACHC_Required_Annual_Training/ComfyUI-Pipeline"


SIDE_COMMENTS: Dict[str, Dict[str, str]] = {
    "M01-S13": {
        "title": "Educator Side Comment: First Contact Signal",
        "note": "Notice what just happened here. Mr. Henderson looked at Marites' hands before he answered. In home health, that micro-pause can signal mistrust, hearing uncertainty, or fear of losing control. A competent clinician slows down, confirms preferences, and documents the interaction signal.",
        "workflow": "Patient encounter communication signal review",
        "risk": "Moderate",
    },
    "M01-S19": {
        "title": "Educator Side Comment: Handoff Safety",
        "note": "This is often missed during busy visits. A rushed handoff is not a personality issue; it is a workflow hazard. Surveyors look for whether critical facts were transferred clearly. If details are dropped, document what is missing and close the loop before you leave.",
        "workflow": "Handoff completeness check",
        "risk": "High",
    },
    "M01-S20": {
        "title": "Educator Side Comment: Interpreter Shortcut Risk",
        "note": "This is where many teams fail. The shortcut sounds practical, but skipping qualified interpretation converts a communication issue into a patient safety issue. A competent clinician identifies language barrier type, activates interpreter workflow, and records interpreter method and ID.",
        "workflow": "Interpreter activation and documentation",
        "risk": "High",
    },
    "M01-S24": {
        "title": "Educator Side Comment: Documentation Gap Escalation Trigger",
        "note": "A surveyor would immediately flag this pattern. If an order changed and the chart did not, you now have a high-risk documentation conflict. Escalation should begin the same visit with objective findings, incident pathway activation, and a clear transfer note.",
        "workflow": "Incident escalation for documentation discrepancy",
        "risk": "High",
    },
    "M01-S26": {
        "title": "Educator Side Comment: Corrective Action Process",
        "note": "Notice the leadership response. Corrective action is not automatic punishment; it is evidence-driven review, documented coaching, and accountability tracking. Competent clinicians cooperate with investigation and preserve objective records instead of private side resolutions.",
        "workflow": "Supervisor coaching and corrective action process",
        "risk": "Moderate",
    },
    "M01-S27": {
        "title": "Educator Side Comment: Retaliation Early Warning",
        "note": "This becomes a legal and workforce safety risk if ignored. Low-volume statements can still be retaliation indicators when tied to a report. The safe move is immediate written documentation and supervisor notification, not personal confrontation.",
        "workflow": "Non-retaliation reporting",
        "risk": "High",
    },
    "M01-S34": {
        "title": "Educator Side Comment: Inclusion to Operations Bridge",
        "note": "Notice how Marites turns observation into system improvement. Inclusion is not symbolic when it changes workflow. A competent clinician proposes process controls that can be audited, adopted, and repeated across the team.",
        "workflow": "Team-level workflow improvement",
        "risk": "Moderate",
    },
    "M01-S38": {
        "title": "Educator Side Comment: What Documentation Must Capture",
        "note": "Key charting checkpoint. The record should capture interpreter ID, consent confirmation, what was taught, what the patient demonstrated in teach-back, and what follow-up was ordered. Without those fields, the visit may feel successful but remain non-defensible.",
        "workflow": "Interpreter and teach-back charting",
        "risk": "High",
    },
    "M01-S41": {
        "title": "Educator Side Comment: Readiness Judgment",
        "note": "Readiness is operational, not emotional. A competent clinician is one whose decisions are consistently sequenced, escalated when needed, and documented with enough specificity that another clinician can safely continue care.",
        "workflow": "Independent shift readiness review",
        "risk": "Moderate",
    },
}


def read_rows() -> List[Dict[str, str]]:
    with CSV_PATH.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def make_side_comment_row(base: Dict[str, str], note_cfg: Dict[str, str], comment_index: int) -> Dict[str, str]:
    row = dict(base)
    row["Scene_ID"] = f"{base['Scene_ID']}-EC{comment_index}"
    row["Slide_Number"] = f"{base['Slide_Number']}-EC{comment_index}"
    row["Scene_Title"] = note_cfg["title"]
    row["Narration_Text"] = note_cfg["note"]
    row["TTS_Pacing_s"] = "32"
    row["Emotion_Tone"] = "Educator coaching"
    row["Transition_In"] = "Soft instructional overlay"
    row["Transition_Out"] = "Return to narrative flow"
    row["ACHC_Topic"] = f"Operational judgment; {base.get('ACHC_Topic', '')}".strip("; ").strip()
    row["Compliance_Category"] = "In-the-moment instruction"
    row["Operational_Workflow"] = note_cfg["workflow"]
    row["Scenario_Type"] = "educator commentary"
    row["Risk_Category"] = note_cfg["risk"]
    row["Escalation_Path"] = row["Escalation_Path"] or "Supervisor pathway when threshold met"
    row["Documentation_Requirement"] = row["Documentation_Requirement"] or "Document observed signal and action taken"
    row["Patient_Safety_Impact"] = "Bridges narrative signal to safety decision making"
    row["Survey_Risk"] = "Moderate"
    row["Slide_Type"] = "educator-commentary"
    row["Challenge_Question"] = ""
    row["Option_A"] = ""
    row["Option_B"] = ""
    row["Option_C"] = ""
    row["Option_D"] = ""
    row["Correct_Answer_Index"] = ""
    row["Correct_Answer_Label"] = ""
    row["Debrief_Text"] = ""
    return row


def build_challenge_map(rows: List[Dict[str, str]]) -> Dict[str, Dict[str, str]]:
    cmap: Dict[str, Dict[str, str]] = {}
    for r in rows:
        if r["Slide_Type"] in {"challenge", "pre-assessment-challenge"}:
            cmap[r["Assessment_Linkage"]] = r
            cmap[r["Scene_ID"]] = r
    return cmap


def option_label(idx: int) -> str:
    return ["A", "B", "C", "D"][idx]


def enrich_debrief(row: Dict[str, str], challenge_row: Dict[str, str]) -> str:
    try:
        correct_idx = int(challenge_row.get("Correct_Answer_Index", "0"))
    except ValueError:
        correct_idx = 0

    options = [
        challenge_row.get("Option_A", "").strip(),
        challenge_row.get("Option_B", "").strip(),
        challenge_row.get("Option_C", "").strip(),
        challenge_row.get("Option_D", "").strip(),
    ]
    correct_text = options[correct_idx] if 0 <= correct_idx < len(options) else "the defensible option"

    wrong_lines = []
    for i, text in enumerate(options):
        if not text or i == correct_idx:
            continue
        wrong_lines.append(
            f"- Option {option_label(i)} risk: {text}. Looks reasonable under pressure, but creates a preventable documentation and escalation gap."
        )

    wrong_section = "\n".join(wrong_lines) if wrong_lines else "- Alternate options create avoidable safety and compliance gaps."

    return (
        f"Operational Case Review\n"
        f"1) Correct decision: Option {option_label(correct_idx)} is correct because it preserves patient safety and workflow integrity under pressure.\n"
        f"2) Why correct: {correct_text}\n"
        f"3) Incorrect option analysis:\n{wrong_section}\n"
        f"4) Patient safety implications: uncontrolled communication barriers or delayed escalation can compound into harm.\n"
        f"5) Workflow implications: sequence matters; right actions in the wrong order still produce risk.\n"
        f"6) Documentation implications: chart objective findings, what was verified, what was escalated, and who received handoff.\n"
        f"7) Escalation implications: trigger supervisor pathway when unresolved clinical, communication, or conduct risk remains.\n"
        f"8) Survey implications: ACHC/CMS reviewers focus on evidence chain, not intent.\n"
        f"9) Operational consequences: unresolved gaps can lead to adverse events, corrective action, and loss of trust.\n"
        f"10) Disciplinary/legal/defensibility: omissions and unsupported charting increase disciplinary risk and weaken legal defense.\n\n"
        f"{row.get('Debrief_Text', '').strip()}"
    ).strip()


def apply_training_layers(rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    challenge_map = build_challenge_map(rows)
    out: List[Dict[str, str]] = []
    comment_idx = 1

    for row in rows:
        slide_type = row["Slide_Type"]
        if slide_type in {"debrief", "pre-assessment-debrief"}:
            key = row.get("Assessment_Linkage", "").strip()
            linked = challenge_map.get(key)
            if linked:
                expanded = enrich_debrief(row, linked)
                row["Debrief_Text"] = expanded
                row["Narration_Text"] = expanded
                row["TTS_Pacing_s"] = str(max(int(row.get("TTS_Pacing_s", "70") or "70"), 95))
                row["Emotion_Tone"] = "Operational case-review coaching"

        out.append(row)

        if row["Scene_ID"] in SIDE_COMMENTS:
            out.append(make_side_comment_row(row, SIDE_COMMENTS[row["Scene_ID"]], comment_idx))
            comment_idx += 1

    return out


def write_csv(rows: List[Dict[str, str]]) -> None:
    headers = list(rows[0].keys())
    with CSV_PATH.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers, quoting=csv.QUOTE_ALL)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(rows: List[Dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "M01 Narration"
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wrap_cols = {"F", "K", "L", "M", "AK"}
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column_letter in wrap_cols:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")

    widths = {
        "A": 18, "B": 13, "C": 12, "D": 42, "E": 40, "F": 80, "G": 12, "H": 30, "I": 24, "J": 24,
        "K": 56, "L": 55, "M": 44, "N": 28, "O": 34, "P": 38, "Q": 18, "R": 22, "S": 36, "T": 20,
        "U": 20, "V": 20, "W": 14, "X": 26, "Y": 30, "Z": 28, "AA": 14, "AB": 16, "AC": 25,
        "AD": 72, "AE": 38, "AF": 38, "AG": 38, "AH": 38, "AI": 12, "AJ": 12, "AK": 84
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    guide = wb.create_sheet("Design Guide")
    guide.append(["Slide_Type", "Meaning"])
    guide.append(["content", "Narrative story delivery"])
    guide.append(["educator-commentary", "Side-comment instructional coaching layer"])
    guide.append(["challenge", "Interactive decision challenge"])
    guide.append(["debrief", "Operational case-review analysis"])
    guide.append(["pre-assessment-intro", "Pre-hook framing card"])
    guide.append(["pre-assessment-challenge", "Pre-hook challenge questions"])
    guide.append(["pre-assessment-debrief", "Pre-hook answer analysis"])

    wb.save(XLSX_PATH)


def to_ts(value: str) -> str:
    return json.dumps(value or "", ensure_ascii=False)


def map_slide_type(row: Dict[str, str]) -> str:
    st = row.get("Slide_Type", "")
    title = row.get("Scene_Title", "")
    if st == "pre-assessment-intro":
        return "pre-assessment-intro"
    if st in {"challenge", "pre-assessment-challenge"}:
        if title.startswith("Final Assessment"):
            return "final-test"
        return "challenge"
    if st in {"debrief", "pre-assessment-debrief"}:
        return "debrief"
    if st == "educator-commentary":
        return "educator-commentary"
    return "content"


def img_seed_from_slide_number(slide_number: str, fallback: int) -> int:
    digits = "".join(ch for ch in slide_number if ch.isdigit())
    return 200 + int(digits[-2:]) if digits else fallback


def write_staging_ts(rows: List[Dict[str, str]]) -> None:
    lines: List[str] = []
    lines.extend(
        [
            "export type SlideType =",
            "  | 'splash'",
            "  | 'objectives'",
            "  | 'pre-assessment-intro'",
            "  | 'content'",
            "  | 'educator-commentary'",
            "  | 'challenge'",
            "  | 'debrief'",
            "  | 'final-test'",
            "  | 'summary'",
            "  | 'certificate';",
            "",
            "export interface M01Slide {",
            "  id: string;",
            "  type: SlideType;",
            "  title: string;",
            "  location?: string;",
            "  narration: string;",
            "  topic?: string;",
            "  policy?: string;",
            "  imgSeed?: number;",
            "  challengeQuestion?: string;",
            "  options?: [string, string, string, string];",
            "  correctAnswer?: number;",
            "  debriefText?: string;",
            "}",
            "",
            "export const M01_SLIDES: M01Slide[] = [",
            "  {",
            "    id: 'M01-SPLASH',",
            "    type: 'splash',",
            "    title: \"Marites' Journey\",",
            "    narration: '',",
            "    topic: 'Cultural Awareness & CLAS Standards',",
            "    policy: 'ACHC-ART-M01 · Annual Training',",
            "    imgSeed: 100,",
            "  },",
            "  {",
            "    id: 'M01-OBJ',",
            "    type: 'objectives',",
            "    title: 'What You Will Learn',",
            "    narration: 'This module combines cinematic story narration, educator side-comments, and challenge/debrief analysis so operational judgment is reinforced continuously in the flow of care.',",
            "    topic: 'Learning Objectives',",
            "    policy: 'ACHC CLAS Domain 1 · HR-TD-001',",
            "    imgSeed: 101,",
            "  },",
        ]
    )

    fallback_seed = 102
    for row in rows:
        stype = map_slide_type(row)
        seed = img_seed_from_slide_number(row.get("Slide_Number", ""), fallback_seed)
        fallback_seed += 1

        lines.append("  {")
        lines.append(f"    id: {to_ts(row.get('Scene_ID', ''))},")
        lines.append(f"    type: {to_ts(stype)},")
        lines.append(f"    title: {to_ts(row.get('Scene_Title', ''))},")
        if row.get("Location"):
            lines.append(f"    location: {to_ts(row.get('Location', ''))},")
        lines.append(f"    narration: {to_ts(row.get('Narration_Text', ''))},")
        if row.get("ACHC_Topic"):
            lines.append(f"    topic: {to_ts(row.get('ACHC_Topic', ''))},")
        if row.get("Policy_Reference"):
            lines.append(f"    policy: {to_ts(row.get('Policy_Reference', ''))},")
        lines.append(f"    imgSeed: {seed},")

        if stype in {"challenge", "final-test"}:
            lines.append(f"    challengeQuestion: {to_ts(row.get('Challenge_Question', ''))},")
            options = [
                row.get("Option_A", ""),
                row.get("Option_B", ""),
                row.get("Option_C", ""),
                row.get("Option_D", ""),
            ]
            lines.append(
                "    options: ["
                + ", ".join(to_ts(opt) for opt in options)
                + "],"
            )
            try:
                cidx = int(row.get("Correct_Answer_Index", "0"))
            except ValueError:
                cidx = 0
            lines.append(f"    correctAnswer: {cidx},")

        if stype == "debrief":
            lines.append(f"    debriefText: {to_ts(row.get('Debrief_Text', ''))},")

        lines.append("  },")

    lines.extend(
        [
            "  {",
            "    id: 'M01-SUMMARY',",
            "    type: 'summary',",
            "    title: 'Module Complete',",
            "    narration: '',",
            "    imgSeed: 998,",
            "  },",
            "  {",
            "    id: 'M01-CERT',",
            "    type: 'certificate',",
            "    title: 'Certificate of Completion',",
            "    narration: '',",
            "    imgSeed: 999,",
            "  },",
            "];",
            "",
            "export const FINAL_TEST_Q_INDICES = M01_SLIDES.reduce<number[]>((acc, s, i) => {",
            "  if (s.type === 'final-test') acc.push(i);",
            "  return acc;",
            "}, []);",
            "",
            "export const PASS_THRESHOLD = 0.8;",
            "",
        ]
    )

    TS_PATH.write_text("\n".join(lines), encoding="utf-8")


def write_pipeline_artifacts() -> None:
    PIPELINE_DIR.mkdir(parents=True, exist_ok=True)

    csv_template = PIPELINE_DIR / "scene_batch_template.csv"
    csv_template.write_text(
        "\n".join(
            [
                "scene_id,slide_number,scene_type,environment,camera_angle,lighting,character_1,character_1_emotion,character_1_wardrobe,character_2,character_2_emotion,character_2_wardrobe,prompt,negative_prompt,reference_pack,seed,aspect_ratio,output_path",
                "M01-S13,13,patient-encounter,mr_henderson_living_room,medium_two_shot,soft_window_daylight,marites,professional_warm,navy_scrubs,mr_henderson,guarded,cardigan,\"Cinematic documentary realism, Marites introduces herself to Mr. Henderson at eye level, subtle tension in first contact, authentic home health environment, natural skin texture, realistic wrinkles, 35mm filmic framing, emotionally grounded.\",\"anime, cartoon, waxy skin, glamor beauty retouch, extra fingers, deformed face, oversaturated HDR\",marites_core_v1+henderson_core_v1,381931,16:9,slides/M01-S13",
                "M01-S24,24,documentation-scene,mr_henderson_home_charting,over_shoulder_close,neutral_interior,marites,focused,navy_scrubs,dana,defensive,scrubs_alt,\"Cinematic healthcare realism, Marites reviewing chart discrepancy while Dana speaks in background, objective tension, realistic EHR charting context, subtle imperfections, natural light rolloff, documentary style.\",\"anime, cartoon, glossy skin, fashion editorial look, plastic lighting\",marites_core_v1+dana_core_v1,918223,16:9,slides/M01-S24",
            ]
        ),
        encoding="utf-8",
    )

    workflow = PIPELINE_DIR / "LMS_Cinematic_FLUX_PuLID_Pipeline.json"
    workflow.write_text(
        json.dumps(
            {
                "pipeline_name": "LMS_Cinematic_FLUX_PuLID_Pipeline",
                "phases": {
                    "phase_1_cast_creation": {
                        "description": "Generate canonical portraits + emotion/wardrobe variants per character.",
                        "characters": ["marites", "aldrin", "don_elena", "ruth", "mr_henderson", "supportive_patient", "hostile_patient", "family_member"],
                        "outputs": ["characters/<name>/front", "characters/<name>/emotion_set", "characters/<name>/wardrobe_set"]
                    },
                    "phase_2_character_locking": {
                        "description": "Multi-reference PuLID conditioning with weighted packs to stabilize identity across lighting/angles.",
                        "reference_strategy": "4 to 8 refs mixed per character by emotion and angle",
                        "weight_guidance": {"base_identity": 0.65, "emotion_variant": 0.2, "lighting_variant": 0.15}
                    },
                    "phase_3_scene_generation": {
                        "description": "CSV-driven scene rendering with fallback compositing for multi-character stability.",
                        "mode_primary": "direct multi-character render",
                        "mode_fallback": ["generate environment plate", "insert character A with masked inpaint", "insert character B with masked inpaint", "lighting harmonization pass", "latent upscale/refine"]
                    }
                },
                "node_blueprint": [
                    {"stage": "csv_ingest", "node": "ComfyUI_Batch_from_CSV", "purpose": "Row-driven generation parameters"},
                    {"stage": "prompt_builder", "node": "Prompt Compose", "purpose": "Build cinematic positive/negative prompts per row"},
                    {"stage": "model_load", "node": "FLUX Checkpoint Loader", "purpose": "Base text-to-image model"},
                    {"stage": "identity", "node": "PuLID Multi-Reference", "purpose": "Character identity lock"},
                    {"stage": "composition", "node": "Regional/Mask Conditioning", "purpose": "Character separation and anti-cloning control"},
                    {"stage": "sampling", "node": "KSampler", "purpose": "Primary latent generation"},
                    {"stage": "inpaint_loop", "node": "Masked Inpaint", "purpose": "Repair faces/hands and preserve character distinction"},
                    {"stage": "upscale", "node": "Latent Upscale + Refiner", "purpose": "Final cinematic quality pass"},
                    {"stage": "save", "node": "Save Image", "purpose": "Route outputs into slides/draft/review/final folders"}
                ],
                "folder_outputs": ["slides", "characters", "environments", "final", "upscaled", "draft", "review"],
                "notes": [
                    "Use realistic skin texture language in prompts; avoid beauty retouch descriptors.",
                    "Prefer 35mm/50mm documentary framing for healthcare realism.",
                    "For unstable multi-character scenes, use hybrid compositing fallback sequence."
                ]
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    runbook = PIPELINE_DIR / "PIPELINE_RUNBOOK.md"
    runbook.write_text(
        "\n".join(
            [
                "# Cinematic LMS Pipeline Runbook",
                "",
                "## Primary Objective",
                "Production workflow for FLUX + PuLID story scenes with character continuity and CSV batch control.",
                "",
                "## Node Architecture",
                "1. **CSV ingest** (`ComfyUI_Batch_from_CSV`) for per-scene parameters.",
                "2. **Character lock** (`PuLID` multi-reference packs with weighted blend).",
                "3. **Scene compose** (regional/masked conditioning to avoid clone artifacts).",
                "4. **Generate** (`KSampler`) with cinematic prompt stack.",
                "5. **Fallback compositing** for unstable multi-character scenes.",
                "6. **Latent upscale/refine** then save into production folders.",
                "",
                "## PuLID Guidance",
                "- Start identity weight around `0.60-0.70`.",
                "- Blend 4-8 references: front, side, low-key light, bright light, neutral expression, emotional expression.",
                "- If waxy face appears: lower identity strength, add texture-rich realism prompt terms, increase denoise control in inpaint pass.",
                "",
                "## Multi-Character Strategy",
                "- Attempt direct render first with separate reference packs.",
                "- If identities collapse, switch to hybrid pipeline:",
                "  1) Generate environment plate",
                "  2) Inpaint Character A",
                "  3) Inpaint Character B",
                "  4) Harmonize shadows/white balance",
                "  5) Run final upscale/refine.",
                "",
                "## Naming Convention",
                "- Characters: `characters/<name>/<name>__<emotion>__<wardrobe>__v01.png`",
                "- Slides: `slides/<scene_id>/<scene_id>__seed<seed>__draft.png`",
                "- Final: `final/<scene_id>__final.png`",
                "",
                "## CSV Columns",
                "`scene_id, slide_number, scene_type, environment, camera_angle, lighting, character_1, character_1_emotion, character_1_wardrobe, character_2, character_2_emotion, character_2_wardrobe, prompt, negative_prompt, reference_pack, seed, aspect_ratio, output_path`",
            ]
        ),
        encoding="utf-8",
    )


def main() -> None:
    rows = read_rows()
    updated_rows = apply_training_layers(rows)
    write_csv(updated_rows)
    write_xlsx(updated_rows)
    write_staging_ts(updated_rows)
    write_pipeline_artifacts()
    print(f"Updated rows: {len(updated_rows)}")
    print(f"CSV: {CSV_PATH}")
    print(f"XLSX: {XLSX_PATH}")
    print(f"TS: {TS_PATH}")
    print(f"Pipeline artifacts: {PIPELINE_DIR}")


if __name__ == "__main__":
    main()
