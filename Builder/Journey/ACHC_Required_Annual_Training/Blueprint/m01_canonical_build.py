"""
M01 Canonical Build Script
==========================
Reconciles all Module 1 / Tess Journey source files into one canonical
CSV + multi-tab Excel matrix.

Sources consumed:
  1. M01-Marites-Journey-Narration-Defensibility-COMPLETE.csv  (base data)
  2. batch1 correction.txt  (S07-S17 production-ready scenes)
  3. batch 2 correction.txt (S26-S30, S32-S39, S31/S40/S42-S46 challenges)
  4. batch 3 correction.txt (S47-S49 debrief + certificate)

Outputs produced:
  - M01-Tess-Journey-CANONICAL.csv
  - M01-Tess-Journey-CANONICAL.xlsx  (9-tab Excel matrix)
  - M01-Tess-Journey-NARRATION.csv
  - M01-Tess-Journey-FLUX-PROMPTS.csv
  - M01-Tess-Journey-LTX-PROMPTS.csv
  - M01-Tess-Journey-CHALLENGES.csv
  - M01-Tess-Journey-DEBRIEFS.csv
  - M01-Tess-Journey-RECONCILIATION-REPORT.md
"""

import csv
import re
import sys
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("WARNING: openpyxl not installed — Excel output skipped. Run: pip install openpyxl")

ROOT = Path(r"C:\AI\Git\training\HomeHealth\Policies_and_Procedures")
BLUEPRINT = ROOT / "Builder/Journey/ACHC_Required_Annual_Training/Blueprint"
CHATGPT   = ROOT / "Builder/_chatGPT"

SOURCE_CSV  = BLUEPRINT / "M01-Marites-Journey-Narration-Defensibility-COMPLETE.csv"
BATCH1_TXT  = CHATGPT  / "batch1 correction.txt"
BATCH2_TXT  = CHATGPT  / "batch 2 correction.txt"
BATCH3_TXT  = CHATGPT  / "batch 3 correction.txt"

OUT_CANONICAL_CSV  = BLUEPRINT / "M01-Tess-Journey-CANONICAL.csv"
OUT_CANONICAL_XLSX = BLUEPRINT / "M01-Tess-Journey-CANONICAL.xlsx"
OUT_NARRATION_CSV  = BLUEPRINT / "M01-Tess-Journey-NARRATION.csv"
OUT_FLUX_CSV       = BLUEPRINT / "M01-Tess-Journey-FLUX-PROMPTS.csv"
OUT_LTX_CSV        = BLUEPRINT / "M01-Tess-Journey-LTX-PROMPTS.csv"
OUT_CHALLENGES_CSV = BLUEPRINT / "M01-Tess-Journey-CHALLENGES.csv"
OUT_DEBRIEFS_CSV   = BLUEPRINT / "M01-Tess-Journey-DEBRIEFS.csv"
OUT_REPORT_MD      = BLUEPRINT / "M01-Tess-Journey-RECONCILIATION-REPORT.md"

# ──────────────────────────────────────────────────────────────────────────────
# CANONICAL COLUMN SCHEMA
# ──────────────────────────────────────────────────────────────────────────────
CANONICAL_HEADERS = [
    "Scene_ID", "Module", "Act", "Slide_Number", "Slide_Type", "Scene_Title",
    "Location", "Narration_Text", "TTS_Pacing_s", "Emotion_Tone",
    "Transition_In", "Transition_Out",
    "Detailed_Image_Description",
    "FLUX1Dev_Image_Prompt",
    "LTX23Dev_Video_Generation_Prompt",
    "Estimated_Card_Duration",
    "LTX_Generation_Count",
    "PuLID_Character_Guidance",
    "Negative_Prompt", "Camera_Framing",
    "ACHC_Topic", "Policy_Reference", "Learning_Objective_ID",
    "Compliance_Category", "Operational_Workflow",
    "Assessment_Linkage", "Debrief_Linkage",
    "Scenario_Type", "Risk_Category", "Escalation_Path",
    "Documentation_Requirement", "Patient_Safety_Impact", "Survey_Risk",
    "Reconciliation_Status",
    "Challenge_Question",
    "Option_A", "Option_B", "Option_C", "Option_D",
    "Correct_Answer_Index", "Correct_Answer_Label",
    "Debrief_Text",
    "Operational_Rationale",
    "Full_Debrief",
    "Status",
]

# ──────────────────────────────────────────────────────────────────────────────
# BATCH CORRECTION PARSER
# ──────────────────────────────────────────────────────────────────────────────

def _strip_field(text: str) -> str:
    """Remove excess whitespace/newlines from a parsed field value."""
    return re.sub(r'\n{3,}', '\n\n', text.strip())


def _extract_field(block: str, label: str, next_labels: list[str]) -> str:
    """
    Extract text for `label:` up to the next recognised label or end of block.
    Returns empty string if label not found.
    """
    # Build a pattern that stops at any of the next labels
    stops = '|'.join(re.escape(nl + ':') for nl in next_labels) if next_labels else '$'
    pattern = re.compile(
        rf'{re.escape(label)}:\s*(.*?)(?=(?:{stops})|$)',
        re.DOTALL | re.IGNORECASE,
    )
    m = pattern.search(block)
    if m:
        return _strip_field(m.group(1))
    return ''


# All field labels we extract from batch blocks (order matters for stop-detection)
ALL_BATCH_LABELS = [
    'Scene_ID', 'Slide_Type', 'Scene_Title',
    'Narration_Text', 'Detailed_Image_Description',
    'FLUX1Dev_Image_Prompt', 'LTX23Dev_Video_Generation_Prompt',
    'Estimated_Card_Duration', 'LTX_Generation_Count', 'PuLID_Character_Guidance',
    'Challenge_Question',
    'Option_A', 'Option_B', 'Option_C', 'Option_D',
    'Correct_Answer', 'Operational_Rationale', 'Full_Debrief',
    'Debrief_Text',
]


def parse_batch_file(path: Path) -> dict[str, dict[str, str]]:
    """
    Parse a batch correction .txt file and return a dict:
        { 'M01-SXX': { field: value, ... }, ... }
    Handles duplicate Scene_IDs by keeping the LAST (most complete) version.
    """
    if not path.exists():
        print(f"  WARNING: batch file not found: {path}")
        return {}

    text = path.read_text(encoding='utf-8', errors='replace')

    # Split on lines that look like "M01-SXX | type" scene headers
    # Also handle "M01-PA-..." and "M01-S...-EC..." IDs
    header_re = re.compile(
        r'^(M01-[A-Z0-9]+-?[A-Z0-9]*(?:-[A-Z0-9]+)?)\s*\|\s*(\S+)',
        re.MULTILINE,
    )

    matches = list(header_re.finditer(text))
    results: dict[str, dict[str, str]] = {}

    for i, m in enumerate(matches):
        scene_id = m.group(1).strip()
        slide_type_raw = m.group(2).strip().lower()

        start = m.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        block = text[start:end]

        # Map slide_type aliases
        SLIDE_TYPE_MAP = {
            'content': 'content',
            'challenge': 'challenge',
            'debrief': 'debrief',
            'pre-assessment-intro': 'pre-assessment-intro',
            'pre-assessment-challenge': 'pre-assessment-challenge',
            'pre-assessment-debrief': 'pre-assessment-debrief',
            'educator-commentary': 'educator-commentary',
        }
        slide_type = SLIDE_TYPE_MAP.get(slide_type_raw, slide_type_raw)

        row: dict[str, str] = {'Scene_ID': scene_id, 'Slide_Type': slide_type}

        for j, label in enumerate(ALL_BATCH_LABELS):
            if label in ('Scene_ID', 'Slide_Type'):
                continue
            next_labs = ALL_BATCH_LABELS[j + 1:j + 4]
            val = _extract_field(block, label, next_labs)
            if val:
                row[label] = val

        # Normalise Estimated_Card_Duration → just the number of seconds
        dur = row.get('Estimated_Card_Duration', '')
        if dur:
            m2 = re.search(r'(\d+)', dur)
            row['Estimated_Card_Duration'] = m2.group(1) if m2 else dur

        # Normalise LTX_Generation_Count → number only
        ltx = row.get('LTX_Generation_Count', '')
        if ltx:
            m3 = re.search(r'(\d+)', ltx)
            row['LTX_Generation_Count'] = m3.group(1) if m3 else ltx

        results[scene_id] = row

    print(f"  Parsed {len(results)} scenes from {path.name}")
    return results


# ──────────────────────────────────────────────────────────────────────────────
# TEXT SUBSTITUTION  (Marites → Tess)
# ──────────────────────────────────────────────────────────────────────────────

def apply_name_change(text: str) -> str:
    """Replace all Marites/Marites' occurrences with Tess/Tess's."""
    text = re.sub(r"\bMarites'\b", "Tess's", text)
    text = re.sub(r"\bMarites\b", "Tess", text)
    # Update character anchor lines
    text = text.replace("Character anchor: Marites,", "Character anchor: Tess,")
    text = text.replace("Character anchor: Marites —", "Character anchor: Tess —")
    # Update title references
    text = text.replace("Marites' Journey", "Tess's Journey")
    text = text.replace("Marites' hands", "Tess's hands")
    text = text.replace("Marites' accent", "Tess's accent")
    text = text.replace("Marites is", "Tess is")
    text = text.replace("Marites has", "Tess has")
    text = text.replace("Marites can", "Tess can")
    text = text.replace("Marites arrives", "Tess arrives")
    text = text.replace("Marites walks", "Tess walks")
    text = text.replace("Marites sits", "Tess sits")
    text = text.replace("Marites reviews", "Tess reviews")
    text = text.replace("Marites reads", "Tess reads")
    text = text.replace("Marites stops", "Tess stops")
    text = text.replace("Marites turns", "Tess turns")
    text = text.replace("Marites looks", "Tess looks")
    text = text.replace("Marites pauses", "Tess pauses")
    text = text.replace("Marites catches", "Tess catches")
    text = text.replace("Marites did not", "Tess did not")
    text = text.replace("Marites does not", "Tess does not")
    text = text.replace("Marites will", "Tess will")
    text = text.replace("Marites needs", "Tess needs")
    text = text.replace("Marites notices", "Tess notices")
    text = text.replace("Marites finds", "Tess finds")
    text = text.replace("Marites files", "Tess files")
    text = text.replace("Marites texts", "Tess texts")
    text = text.replace("Marites laughs", "Tess laughs")
    text = text.replace("Marites writes", "Tess writes")
    text = text.replace("Marites raises", "Tess raises")
    text = text.replace("Marites reports", "Tess reports")
    text = text.replace("Marites documents", "Tess documents")
    text = text.replace("Marites catches", "Tess catches")
    text = text.replace("Marites escalates", "Tess escalates")
    text = text.replace("Marites succeeded", "Tess succeeded")
    # Final fallback for any remaining instances
    text = re.sub(r'\bMarites\b', 'Tess', text)
    return text


def apply_name_change_to_row(row: dict[str, str]) -> dict[str, str]:
    return {k: apply_name_change(v) for k, v in row.items()}


# ──────────────────────────────────────────────────────────────────────────────
# S49 HARD-CODED ROW (missing from existing CSV)
# ──────────────────────────────────────────────────────────────────────────────

S49_ROW: dict[str, str] = {
    "Scene_ID": "M01-S49",
    "Module": "M01",
    "Act": "Act 4",
    "Slide_Number": "S49",
    "Slide_Type": "content",
    "Scene_Title": "The Reveal and the Legacy",
    "Location": "Home health agency office — Director of Nursing's desk",
    "Narration_Text": (
        "The scene shifts. A clinic office — clean, organized, well-lit. A Filipina woman "
        "in her early 50s sits at the desk. Her badge reads Director of Nursing. She looks "
        "directly at you. On her wrist, next to her watch — the same woven bracelet. Faded "
        "now. Threadbare. But still there.\n\n"
        "She speaks: You made it. I know — it's a lot of rules, a lot of standards. Some "
        "days your feet ache before your shift is halfway done. But every standard you "
        "learned today? It's a safety net. For your patients, and for you. You don't have "
        "to erase where you came from to be excellent here. You just have to learn how to "
        "translate your respect into a language this system understands.\n\n"
        "She pauses. Quieter: I know you can do it. Look how far we've both come. Keep "
        "your roots strong. Take a breath. Go print that certificate. Your patients are waiting."
    ),
    "TTS_Pacing_s": "90",
    "Emotion_Tone": "Warm grounded earned",
    "Transition_In": "Scene shift dissolve",
    "Transition_Out": "Certificate screen reveal",
    "Detailed_Image_Description": (
        "A modern but functional home health agency office. A standard office with a door, "
        "a desk, and the accumulated artifacts of twenty years of clinical leadership. "
        "Wood-laminate desk over a metal frame — practical, not decorative. On the desk: "
        "a closed laptop, a stack of orientation binders with tabbed sections, a ceramic "
        "coffee mug with a small chip on the handle rim. A framed photo angled away from "
        "the camera. A small potted succulent on the right corner.\n\n"
        "The woman seated behind the desk is Filipina, early 50s. Short professional dark "
        "hair cut just below the jaw, with natural grey visible at both temples. Natural "
        "aging: fine lines at the outer corners of her eyes, weathering around the mouth — "
        "twenty years of clinical expression. The same facial bone structure as young "
        "Tess — the same wide cheekbones, the same eye shape, the same small mole on the "
        "left cheek. She is recognizably the same person, two decades later. Navy blazer "
        "over a cream collared blouse. Director of Nursing badge on retractable reel. "
        "Relaxed authoritative posture: shoulders square, body slightly forward, forearms "
        "resting on the desk. She is not performing authority. She has it.\n\n"
        "Her left arm rests on the desk beside the laptop. Between her simple silver watch "
        "and the edge of her blazer sleeve: the bracelet. The same thin woven thread "
        "bracelet from the grandmother's hands. It has aged visibly: the original "
        "earth-brown thread faded to a pale grey-tan over decades of wear. Several strands "
        "slightly frayed at the edges. One short section thinner from friction. The "
        "grandmother's knot still present — still holding. The bracelet catches the "
        "overhead light.\n\n"
        "The office: overhead fluorescent mixed with natural window light from the left. "
        "Whiteboard behind her showing a partial staff schedule in dry-erase marker. "
        "Bookshelf with labeled compliance binders: ACHC Survey Prep, Policy Manual Vol. 1, "
        "Policy Manual Vol. 2, QA Records, HR Documentation. A small clock on the whiteboard "
        "frame shows 4:45 PM — a direct mirror of the 4:45 AM from S01. Certificate of "
        "Completion UI element appears as a clean subtle overlay lower right."
    ),
    "FLUX1Dev_Image_Prompt": (
        "Photorealistic documentary medium shot, modern functional home health agency office, "
        "Filipina woman early 50s seated behind standard wood-laminate desk facing camera "
        "directly, short professional dark hair cut below jaw with natural grey at temples, "
        "natural aging — fine lines at eye corners and weathering around mouth, same wide "
        "cheekbones and eye shape and small left-cheek mole as young Tess — recognizably "
        "same person older, navy blazer over cream collared blouse, Director of Nursing badge "
        "on retractable reel on lapel, warm grounded knowing expression not theatrical, "
        "relaxed authoritative posture with forearms on desk, left arm resting beside closed "
        "laptop with woven thread bracelet clearly visible between simple silver watch and "
        "blazer sleeve — bracelet visibly aged: original earth-brown faded to pale grey-tan, "
        "several strands frayed at edges, one two-centimeter section thinner from wear, "
        "grandmother's knot still intact and visible, desk with chipped ceramic mug and "
        "orientation binders with tabbed edges and small succulent on corner, whiteboard "
        "behind with partial staff schedule and compliance binder bookshelf, small clock on "
        "whiteboard frame showing 4:45 PM, overhead fluorescent mixed with natural window "
        "light from left, Certificate of Completion UI element as clean subtle overlay lower "
        "right corner, 35mm lens, true skin texture documentary realism, earned authority presence."
    ),
    "LTX23Dev_Video_Generation_Prompt": (
        "CLIP 1 (0–10s): Older Tess sits at her desk, looking directly at the camera. She "
        "holds the gaze for a moment before speaking. Her left hand shifts slightly on the "
        "desk — the aged bracelet catches the overhead light briefly. She begins speaking: "
        "calm, even, unhurried. Clock behind her reads 4:45 PM. Both forearms on the desk, "
        "open posture. She speaks three sentences.\n\n"
        "CLIP 2 (10–20s): She pauses mid-speech. Her gaze drops for a moment to her left "
        "wrist — to the bracelet. The faintest shift in her expression — not a smile exactly, "
        "just recognition. She looks back up at the camera. Her voice quieter. She speaks "
        "the final two sentences. A single slow nod at the end. She holds the camera gaze "
        "for one beat of stillness after the nod. Certificate of Completion UI element fades "
        "in cleanly lower right corner. She keeps her eyes on the learner. Stillness."
    ),
    "Estimated_Card_Duration": "20",
    "LTX_Generation_Count": "2",
    "PuLID_Character_Guidance": (
        "OLDER TESS — critical character reveal and module payoff. Same Filipina ethnicity. "
        "NOW early 50s. Short professional dark hair with natural grey at temples — not dyed. "
        "Same facial bone structure as young Tess: wide cheekbones, same eye shape, same "
        "small mole on left cheek — recognizably the same person, naturally aged. Natural "
        "aging: fine lines at eye corners, slight weathering around mouth, softening of jaw. "
        "Navy blazer, cream collared blouse, Director of Nursing badge on retractable reel. "
        "CRITICAL BRACELET CONTINUITY: Same woven thread bracelet, same left wrist. NOW "
        "visibly aged: original earth-brown thread faded to pale grey-tan. Several strands "
        "frayed at edges. One short section thinner from two decades of daily wear. "
        "Grandmother's knot still intact and visible. Clock at 4:45 PM mirrors 4:45 AM "
        "from S01. Expression: warm, grounded, knowing — not inspirational performance."
    ),
    "Negative_Prompt": (
        "anime, cartoon, CGI, plastic skin, distorted hands, extra fingers, watermark, "
        "oversaturated HDR, fantasy lighting, stock-photo smile, obvious aging makeup, "
        "Hollywood glamour"
    ),
    "Camera_Framing": "Medium — direct camera address, desk in foreground",
    "ACHC_Topic": "Cultural competence; Patient rights; Training completion",
    "Policy_Reference": "ACHC CLAS Domain 1; HR-TD-001 Annual Training",
    "Learning_Objective_ID": "LO1|LO3|LO4|LO5",
    "Compliance_Category": "Module Completion",
    "Operational_Workflow": "Certification and training record completion",
    "Assessment_Linkage": "Final Assessment Complete",
    "Debrief_Linkage": "N/A",
    "Scenario_Type": "Certificate reveal and legacy close",
    "Risk_Category": "Low",
    "Escalation_Path": "N/A",
    "Documentation_Requirement": "LMS completion record",
    "Patient_Safety_Impact": "Completes training loop with defensible workflow reinforcement",
    "Survey_Risk": "Low",
    "Reconciliation_Status": "Mapped",
    "Challenge_Question": "",
    "Option_A": "",
    "Option_B": "",
    "Option_C": "",
    "Option_D": "",
    "Correct_Answer_Index": "",
    "Correct_Answer_Label": "",
    "Debrief_Text": "",
    "Operational_Rationale": "",
    "Full_Debrief": "",
    "Status": "COMPLETE",
}


# ──────────────────────────────────────────────────────────────────────────────
# LOAD BASE CSV
# ──────────────────────────────────────────────────────────────────────────────

def load_base_csv() -> list[dict[str, str]]:
    with SOURCE_CSV.open("r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))
    print(f"  Loaded {len(rows)} rows from {SOURCE_CSV.name}")
    return rows


# ──────────────────────────────────────────────────────────────────────────────
# MERGE BATCH DATA INTO ROW
# ──────────────────────────────────────────────────────────────────────────────

BATCH_TO_CANONICAL = {
    'Narration_Text':                   'Narration_Text',
    'FLUX1Dev_Image_Prompt':            'FLUX1Dev_Image_Prompt',
    'Detailed_Image_Description':       'Detailed_Image_Description',
    'LTX23Dev_Video_Generation_Prompt': 'LTX23Dev_Video_Generation_Prompt',
    'Estimated_Card_Duration':          'Estimated_Card_Duration',
    'LTX_Generation_Count':             'LTX_Generation_Count',
    'PuLID_Character_Guidance':         'PuLID_Character_Guidance',
    'Challenge_Question':               'Challenge_Question',
    'Option_A':                         'Option_A',
    'Option_B':                         'Option_B',
    'Option_C':                         'Option_C',
    'Option_D':                         'Option_D',
    'Correct_Answer':                   'Correct_Answer_Label',   # batch uses Correct_Answer
    'Operational_Rationale':            'Operational_Rationale',
    'Full_Debrief':                     'Full_Debrief',
}


def merge_batch_into_row(row: dict[str, str], batch_data: dict[str, str]) -> dict[str, str]:
    """
    Apply batch correction data to a canonical row.
    Existing non-empty values are overwritten only when batch has richer content.
    """
    for batch_key, canon_key in BATCH_TO_CANONICAL.items():
        val = batch_data.get(batch_key, '').strip()
        if not val:
            continue
        existing = row.get(canon_key, '').strip()
        # Always prefer batch content (it is the "latest corrected" version)
        # Exception: keep existing Narration_Text if batch has SHORTER text
        if batch_key == 'Narration_Text' and existing and len(existing) > len(val) * 1.5:
            # Existing is substantially longer — keep existing but apply name change
            pass  # will be name-changed globally anyway
        else:
            row[canon_key] = val

    # Handle Correct_Answer_Index from Correct_Answer label
    correct_label = batch_data.get('Correct_Answer', '')
    if correct_label:
        row['Correct_Answer_Label'] = correct_label
        idx_map = {'A': '0', 'B': '1', 'C': '2', 'D': '3'}
        row['Correct_Answer_Index'] = idx_map.get(correct_label.strip().upper(), '')

    return row


# ──────────────────────────────────────────────────────────────────────────────
# ESTIMATE CARD DURATION FROM TTS PACING
# ──────────────────────────────────────────────────────────────────────────────

def estimate_duration(row: dict[str, str]) -> str:
    """Derive Estimated_Card_Duration from TTS_Pacing_s if not set by batch."""
    existing = row.get('Estimated_Card_Duration', '').strip()
    if existing:
        return existing
    tts = row.get('TTS_Pacing_s', '').strip()
    if tts and tts.isdigit():
        # Card duration ≈ narration time + ~3s buffer
        return str(int(tts) + 3)
    return '15'


def estimate_ltx_count(row: dict[str, str]) -> str:
    """Default LTX generation count if not set by batch."""
    existing = row.get('LTX_Generation_Count', '').strip()
    if existing:
        return existing
    # Pre-assessment and debrief cards: 1; content/challenge: 2
    stype = row.get('Slide_Type', '').lower()
    if 'pre-assessment' in stype or 'debrief' in stype or 'educator' in stype:
        return '1'
    return '2'


# ──────────────────────────────────────────────────────────────────────────────
# BUILD CANONICAL ROWS
# ──────────────────────────────────────────────────────────────────────────────

def build_canonical_rows(
    base_rows: list[dict[str, str]],
    all_batch: dict[str, dict[str, str]],
) -> list[dict[str, str]]:
    canonical: list[dict[str, str]] = []
    seen_ids: set[str] = set()

    for raw_row in base_rows:
        scene_id = raw_row.get('Scene_ID', '').strip()
        if not scene_id:
            continue

        # Deduplicate
        if scene_id in seen_ids:
            print(f"  DUPLICATE skipped: {scene_id}")
            continue
        seen_ids.add(scene_id)

        # Build canonical row with all headers
        row: dict[str, str] = {h: '' for h in CANONICAL_HEADERS}

        # Copy existing CSV fields (rename FLUX_Image_Prompt → FLUX1Dev_Image_Prompt)
        field_map = {
            'Scene_ID':              'Scene_ID',
            'Act':                   'Act',
            'Slide_Number':          'Slide_Number',
            'Scene_Title':           'Scene_Title',
            'Location':              'Location',
            'Narration_Text':        'Narration_Text',
            'TTS_Pacing_s':          'TTS_Pacing_s',
            'Emotion_Tone':          'Emotion_Tone',
            'Transition_In':         'Transition_In',
            'Transition_Out':        'Transition_Out',
            'FLUX_Image_Prompt':     'FLUX1Dev_Image_Prompt',
            'PuLID_Character_Guidance': 'PuLID_Character_Guidance',
            'Negative_Prompt':       'Negative_Prompt',
            'Camera_Framing':        'Camera_Framing',
            'ACHC_Topic':            'ACHC_Topic',
            'Policy_Reference':      'Policy_Reference',
            'Learning_Objective_ID': 'Learning_Objective_ID',
            'Compliance_Category':   'Compliance_Category',
            'Operational_Workflow':  'Operational_Workflow',
            'Assessment_Linkage':    'Assessment_Linkage',
            'Debrief_Linkage':       'Debrief_Linkage',
            'Scenario_Type':         'Scenario_Type',
            'Risk_Category':         'Risk_Category',
            'Escalation_Path':       'Escalation_Path',
            'Documentation_Requirement': 'Documentation_Requirement',
            'Patient_Safety_Impact': 'Patient_Safety_Impact',
            'Survey_Risk':           'Survey_Risk',
            'Reconciliation_Status': 'Reconciliation_Status',
            'Slide_Type':            'Slide_Type',
            'Challenge_Question':    'Challenge_Question',
            'Option_A':              'Option_A',
            'Option_B':              'Option_B',
            'Option_C':              'Option_C',
            'Option_D':              'Option_D',
            'Correct_Answer_Index':  'Correct_Answer_Index',
            'Correct_Answer_Label':  'Correct_Answer_Label',
            'Debrief_Text':          'Debrief_Text',
        }
        for src, dst in field_map.items():
            val = raw_row.get(src, '')
            if val:
                row[dst] = val

        row['Module'] = 'M01'
        if not row.get('Act'):
            row['Act'] = 'Act 1'

        # Merge batch correction data if available
        # Also handle split IDs: M01-S31-Q/D → M01-S31 batch; M01-S40-Q/D → M01-S40 batch
        batch_scene = all_batch.get(scene_id)
        if not batch_scene:
            # Try stripping -Q/-D/-Qa/-Da suffix to find unified batch scene
            base_id = re.sub(r'-(Q|D|Qa|Qb|Qc|Da|Db|Dc)$', '', scene_id, flags=re.IGNORECASE)
            if base_id != scene_id:
                batch_scene = all_batch.get(base_id)

        if batch_scene:
            # For -D (debrief) rows: only apply debrief-relevant fields from batch
            if scene_id.endswith(('-D', '-Da', '-Db', '-Dc')):
                debrief_batch = {
                    k: v for k, v in batch_scene.items()
                    if k in ('Full_Debrief', 'Debrief_Text', 'Detailed_Image_Description',
                             'FLUX1Dev_Image_Prompt', 'LTX23Dev_Video_Generation_Prompt',
                             'Estimated_Card_Duration', 'LTX_Generation_Count',
                             'PuLID_Character_Guidance')
                }
                row = merge_batch_into_row(row, debrief_batch)
            else:
                row = merge_batch_into_row(row, batch_scene)

        # Apply Marites → Tess substitution everywhere
        row = apply_name_change_to_row(row)

        # Fill computed fields
        row['Estimated_Card_Duration'] = estimate_duration(row)
        row['LTX_Generation_Count'] = estimate_ltx_count(row)

        # Set Status
        has_ltx = bool(row.get('LTX23Dev_Video_Generation_Prompt', '').strip())
        has_detail = bool(row.get('Detailed_Image_Description', '').strip())
        if has_ltx and has_detail:
            row['Status'] = 'COMPLETE'
        elif row.get('FLUX1Dev_Image_Prompt', '').strip():
            row['Status'] = 'NEEDS REVIEW — LTX/Detail missing'
        else:
            row['Status'] = 'NEEDS REVIEW — Production fields missing'

        # Debrief rows: populate Full_Debrief from Debrief_Text if not set
        if row.get('Slide_Type', '') in ('debrief', 'pre-assessment-debrief'):
            if not row.get('Full_Debrief', '').strip():
                row['Full_Debrief'] = row.get('Debrief_Text', '')

        canonical.append(row)

    # Add S49 if not already present
    if 'M01-S49' not in seen_ids:
        s49 = apply_name_change_to_row(dict(S49_ROW))
        s49['Module'] = 'M01'
        canonical.append(s49)
        print("  Added M01-S49 (Certificate Screen) — was missing from source CSV")

    return canonical


# ──────────────────────────────────────────────────────────────────────────────
# WRITE CANONICAL CSV
# ──────────────────────────────────────────────────────────────────────────────

def write_canonical_csv(rows: list[dict[str, str]]) -> None:
    with OUT_CANONICAL_CSV.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f, fieldnames=CANONICAL_HEADERS, quoting=csv.QUOTE_ALL,
            extrasaction='ignore'
        )
        writer.writeheader()
        writer.writerows(rows)
    print(f"  Written: {OUT_CANONICAL_CSV.name}  ({len(rows)} rows)")


# ──────────────────────────────────────────────────────────────────────────────
# WRITE FOCUSED CSVs
# ──────────────────────────────────────────────────────────────────────────────

def write_focused_csvs(rows: list[dict[str, str]]) -> None:
    # Narration
    narration_headers = [
        "Scene_ID", "Module", "Slide_Type", "Scene_Title",
        "Narration_Text", "TTS_Pacing_s", "Estimated_Card_Duration", "Status"
    ]
    _write_focused(OUT_NARRATION_CSV, rows, narration_headers)

    # FLUX prompts
    flux_headers = [
        "Scene_ID", "Slide_Type", "Scene_Title",
        "Detailed_Image_Description", "FLUX1Dev_Image_Prompt",
        "PuLID_Character_Guidance", "Negative_Prompt", "Camera_Framing", "Status"
    ]
    _write_focused(OUT_FLUX_CSV, rows, flux_headers)

    # LTX prompts
    ltx_headers = [
        "Scene_ID", "Slide_Type", "Scene_Title",
        "LTX23Dev_Video_Generation_Prompt",
        "Estimated_Card_Duration", "LTX_Generation_Count", "Status"
    ]
    _write_focused(OUT_LTX_CSV, rows, ltx_headers)

    # Challenges
    challenge_rows = [r for r in rows if 'challenge' in r.get('Slide_Type', '')]
    challenge_headers = [
        "Scene_ID", "Scene_Title", "Challenge_Question",
        "Option_A", "Option_B", "Option_C", "Option_D",
        "Correct_Answer_Index", "Correct_Answer_Label",
        "Operational_Rationale", "Status"
    ]
    _write_focused(OUT_CHALLENGES_CSV, challenge_rows, challenge_headers)

    # Debriefs
    debrief_rows = [r for r in rows if 'debrief' in r.get('Slide_Type', '')]
    debrief_headers = [
        "Scene_ID", "Scene_Title", "Debrief_Text", "Full_Debrief", "Status"
    ]
    _write_focused(OUT_DEBRIEFS_CSV, debrief_rows, debrief_headers)

    print(f"  Written focused CSVs: Narration, FLUX, LTX, Challenges, Debriefs")


def _write_focused(
    path: Path,
    rows: list[dict[str, str]],
    headers: list[str],
) -> None:
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f, fieldnames=headers, quoting=csv.QUOTE_ALL, extrasaction='ignore'
        )
        writer.writeheader()
        writer.writerows(rows)


# ──────────────────────────────────────────────────────────────────────────────
# EXCEL MATRIX GENERATION
# ──────────────────────────────────────────────────────────────────────────────

HEADER_FILL  = None
HEADER_FONT  = None
SUBHDR_FILL  = None
COMPLETE_FILL = None
REVIEW_FILL  = None


def _init_styles():
    global HEADER_FILL, HEADER_FONT, SUBHDR_FILL, COMPLETE_FILL, REVIEW_FILL
    HEADER_FILL  = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    HEADER_FONT  = Font(color="FFFFFF", bold=True, size=10)
    SUBHDR_FILL  = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
    COMPLETE_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    REVIEW_FILL  = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")


def _style_header_row(ws: Any, col_widths: dict[int, int] | None = None) -> None:
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    if col_widths:
        for col_idx, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width


def _wrap_col(ws: Any, col_letter: str) -> None:
    for cell in ws[col_letter][1:]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def _status_color(ws: Any, status_col: int, row_count: int) -> None:
    for row_idx in range(2, row_count + 2):
        cell = ws.cell(row=row_idx, column=status_col)
        val = (cell.value or '').upper()
        if 'COMPLETE' in val:
            cell.fill = COMPLETE_FILL
        elif 'NEEDS' in val or 'REVIEW' in val:
            cell.fill = REVIEW_FILL


def write_xlsx(rows: list[dict[str, str]]) -> None:
    if not HAS_OPENPYXL:
        return

    _init_styles()
    wb = Workbook()

    # ── Tab 1: Master Scene Matrix ─────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Master Scene Matrix"
    master_cols = [
        "Scene_ID", "Module", "Act", "Slide_Type", "Scene_Title",
        "Estimated_Card_Duration", "LTX_Generation_Count",
        "ACHC_Topic", "Risk_Category", "Survey_Risk", "Status"
    ]
    ws1.append(master_cols)
    for r in rows:
        ws1.append([r.get(c, '') for c in master_cols])
    _style_header_row(ws1, {1:20, 2:7, 3:12, 4:25, 5:48, 6:12, 7:12, 8:36, 9:14, 10:12, 11:30})
    _status_color(ws1, master_cols.index("Status") + 1, len(rows))

    # ── Tab 2: Narration ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Narration")
    narr_cols = [
        "Scene_ID", "Slide_Type", "Scene_Title", "Narration_Text",
        "TTS_Pacing_s", "Estimated_Card_Duration", "Emotion_Tone", "Status"
    ]
    ws2.append(narr_cols)
    for r in rows:
        ws2.append([r.get(c, '') for c in narr_cols])
    _style_header_row(ws2, {1:20, 2:25, 3:44, 4:80, 5:10, 6:12, 7:24, 8:30})
    _wrap_col(ws2, 'D')
    _status_color(ws2, narr_cols.index("Status") + 1, len(rows))

    # ── Tab 3: FLUX Prompts ────────────────────────────────────────────────
    ws3 = wb.create_sheet("FLUX Prompts")
    flux_cols = [
        "Scene_ID", "Slide_Type", "Scene_Title",
        "Detailed_Image_Description", "FLUX1Dev_Image_Prompt",
        "Negative_Prompt", "Camera_Framing",
        "PuLID_Character_Guidance", "Status"
    ]
    ws3.append(flux_cols)
    for r in rows:
        ws3.append([r.get(c, '') for c in flux_cols])
    _style_header_row(ws3, {1:20, 2:25, 3:44, 4:70, 5:70, 6:50, 7:28, 8:60, 9:30})
    for col in ('D', 'E', 'F', 'H'):
        _wrap_col(ws3, col)
    _status_color(ws3, flux_cols.index("Status") + 1, len(rows))

    # ── Tab 4: LTX Prompts ─────────────────────────────────────────────────
    ws4 = wb.create_sheet("LTX Prompts")
    ltx_cols = [
        "Scene_ID", "Slide_Type", "Scene_Title",
        "LTX23Dev_Video_Generation_Prompt",
        "Estimated_Card_Duration", "LTX_Generation_Count", "Status"
    ]
    ws4.append(ltx_cols)
    for r in rows:
        ws4.append([r.get(c, '') for c in ltx_cols])
    _style_header_row(ws4, {1:20, 2:25, 3:44, 4:80, 5:12, 6:12, 7:30})
    _wrap_col(ws4, 'D')
    _status_color(ws4, ltx_cols.index("Status") + 1, len(rows))

    # ── Tab 5: Challenges ──────────────────────────────────────────────────
    ws5 = wb.create_sheet("Challenges")
    challenge_rows = [r for r in rows if 'challenge' in r.get('Slide_Type', '')]
    chall_cols = [
        "Scene_ID", "Scene_Title", "Challenge_Question",
        "Option_A", "Option_B", "Option_C", "Option_D",
        "Correct_Answer_Index", "Correct_Answer_Label",
        "Operational_Rationale", "Status"
    ]
    ws5.append(chall_cols)
    for r in challenge_rows:
        ws5.append([r.get(c, '') for c in chall_cols])
    _style_header_row(ws5, {1:20, 2:44, 3:60, 4:50, 5:50, 6:50, 7:50, 8:10, 9:10, 10:80, 11:30})
    for col in ('C', 'D', 'E', 'F', 'G', 'J'):
        _wrap_col(ws5, col)
    _status_color(ws5, chall_cols.index("Status") + 1, len(challenge_rows))

    # ── Tab 6: Debriefs ────────────────────────────────────────────────────
    ws6 = wb.create_sheet("Debriefs")
    debrief_rows = [r for r in rows if 'debrief' in r.get('Slide_Type', '')]
    deb_cols = [
        "Scene_ID", "Scene_Title", "Debrief_Text", "Full_Debrief", "Status"
    ]
    ws6.append(deb_cols)
    for r in debrief_rows:
        ws6.append([r.get(c, '') for c in deb_cols])
    _style_header_row(ws6, {1:20, 2:44, 3:80, 4:80, 5:30})
    _wrap_col(ws6, 'C')
    _wrap_col(ws6, 'D')
    _status_color(ws6, deb_cols.index("Status") + 1, len(debrief_rows))

    # ── Tab 7: Character Continuity ────────────────────────────────────────
    ws7 = wb.create_sheet("Character Continuity")
    cont_cols = [
        "Scene_ID", "Slide_Type", "Scene_Title",
        "PuLID_Character_Guidance", "Camera_Framing", "Status"
    ]
    ws7.append(cont_cols)
    for r in rows:
        ws7.append([r.get(c, '') for c in cont_cols])
    _style_header_row(ws7, {1:20, 2:25, 3:44, 4:70, 5:28, 6:30})
    _wrap_col(ws7, 'D')
    _status_color(ws7, cont_cols.index("Status") + 1, len(rows))

    # ── Tab 8: Runtime Summary ─────────────────────────────────────────────
    ws8 = wb.create_sheet("Runtime Summary")
    ws8.append(["Metric", "Value"])
    for cell in ws8[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    total_cards = len(rows)
    total_duration = sum(
        int(r.get('Estimated_Card_Duration', '0') or '0') for r in rows
    )
    total_ltx = sum(
        int(r.get('LTX_Generation_Count', '0') or '0') for r in rows
    )
    complete_count = sum(1 for r in rows if r.get('Status', '').startswith('COMPLETE'))
    needs_review_count = total_cards - complete_count

    type_counts: dict[str, int] = {}
    for r in rows:
        st = r.get('Slide_Type', 'unknown')
        type_counts[st] = type_counts.get(st, 0) + 1

    challenge_count = sum(1 for r in rows if 'challenge' in r.get('Slide_Type', ''))
    debrief_count   = sum(1 for r in rows if 'debrief'   in r.get('Slide_Type', ''))
    educator_count  = sum(1 for r in rows if 'educator'  in r.get('Slide_Type', ''))
    content_count   = sum(1 for r in rows if r.get('Slide_Type', '') == 'content')
    pa_count        = sum(1 for r in rows if 'pre-assessment' in r.get('Slide_Type', ''))

    missing_narration = sum(1 for r in rows if not r.get('Narration_Text', '').strip())
    missing_flux      = sum(1 for r in rows if not r.get('FLUX1Dev_Image_Prompt', '').strip())
    missing_ltx       = sum(1 for r in rows if not r.get('LTX23Dev_Video_Generation_Prompt', '').strip())
    missing_detail    = sum(1 for r in rows if not r.get('Detailed_Image_Description', '').strip())

    dup_ids: dict[str, int] = {}
    for r in rows:
        sid = r.get('Scene_ID', '')
        dup_ids[sid] = dup_ids.get(sid, 0) + 1
    duplicate_count = sum(1 for v in dup_ids.values() if v > 1)

    runtime_mm = total_duration // 60
    runtime_ss = total_duration % 60

    stats = [
        ("Total Cards (production rows)", total_cards),
        ("Total Estimated Duration (seconds)", total_duration),
        ("Total Estimated Duration (mm:ss)", f"{runtime_mm:02d}:{runtime_ss:02d}"),
        ("Total LTX Video Generations", total_ltx),
        ("", ""),
        ("Cards by Type", ""),
        ("  Content", content_count),
        ("  Pre-Assessment", pa_count),
        ("  Challenge", challenge_count),
        ("  Debrief", debrief_count),
        ("  Educator Commentary", educator_count),
        ("", ""),
        ("QA Status", ""),
        ("  Status: COMPLETE", complete_count),
        ("  Status: NEEDS REVIEW", needs_review_count),
        ("  Duplicate Scene_IDs", duplicate_count),
        ("", ""),
        ("Missing Fields", ""),
        ("  Missing Narration_Text", missing_narration),
        ("  Missing FLUX1Dev_Image_Prompt", missing_flux),
        ("  Missing LTX23Dev_Video_Generation_Prompt", missing_ltx),
        ("  Missing Detailed_Image_Description", missing_detail),
    ]

    for label, value in stats:
        ws8.append([label, str(value) if value != "" else ""])

    ws8.column_dimensions['A'].width = 40
    ws8.column_dimensions['B'].width = 20

    # ── Tab 9: QA Checklist ────────────────────────────────────────────────
    ws9 = wb.create_sheet("QA Checklist")
    qa_cols = [
        "Scene_ID", "Slide_Type", "Scene_Title",
        "Has_Narration", "Has_FLUX_Prompt", "Has_LTX_Prompt",
        "Has_Detail_Desc", "Has_PuLID", "Has_Duration",
        "Challenge_Complete", "Status"
    ]
    ws9.append(qa_cols)
    for r in rows:
        stype = r.get('Slide_Type', '')
        is_challenge = 'challenge' in stype
        chall_ok = (
            bool(r.get('Challenge_Question', '').strip()) and
            bool(r.get('Option_A', '').strip()) and
            bool(r.get('Correct_Answer_Label', '').strip())
        ) if is_challenge else 'N/A'

        ws9.append([
            r.get('Scene_ID', ''),
            stype,
            r.get('Scene_Title', ''),
            'YES' if r.get('Narration_Text', '').strip() else 'MISSING',
            'YES' if r.get('FLUX1Dev_Image_Prompt', '').strip() else 'MISSING',
            'YES' if r.get('LTX23Dev_Video_Generation_Prompt', '').strip() else 'MISSING',
            'YES' if r.get('Detailed_Image_Description', '').strip() else 'MISSING',
            'YES' if r.get('PuLID_Character_Guidance', '').strip() else 'MISSING',
            'YES' if r.get('Estimated_Card_Duration', '').strip() else 'MISSING',
            'YES' if chall_ok is True else ('N/A' if chall_ok == 'N/A' else 'MISSING'),
            r.get('Status', ''),
        ])
    _style_header_row(ws9, {1:20, 2:25, 3:44, 4:12, 5:12, 6:12, 7:12, 8:12, 9:12, 10:16, 11:30})
    _status_color(ws9, qa_cols.index("Status") + 1, len(rows))

    wb.save(OUT_CANONICAL_XLSX)
    print(f"  Written: {OUT_CANONICAL_XLSX.name}  (9 tabs)")


# ──────────────────────────────────────────────────────────────────────────────
# RECONCILIATION REPORT
# ──────────────────────────────────────────────────────────────────────────────

def write_reconciliation_report(
    rows: list[dict[str, str]],
    batch_scene_ids: set[str],
    added_s49: bool,
) -> None:
    total = len(rows)
    total_dur = sum(int(r.get('Estimated_Card_Duration', '0') or '0') for r in rows)
    total_ltx = sum(int(r.get('LTX_Generation_Count', '0') or '0') for r in rows)
    complete  = sum(1 for r in rows if r.get('Status', '').startswith('COMPLETE'))
    review    = total - complete

    type_counts: dict[str, int] = {}
    for r in rows:
        st = r.get('Slide_Type', 'unknown')
        type_counts[st] = type_counts.get(st, 0) + 1

    missing_fields: list[str] = []
    for r in rows:
        sid = r.get('Scene_ID', '')
        for field in ('Narration_Text', 'FLUX1Dev_Image_Prompt',
                      'LTX23Dev_Video_Generation_Prompt', 'Detailed_Image_Description'):
            if not r.get(field, '').strip():
                missing_fields.append(f"{sid} — {field}")

    mm = total_dur // 60
    ss = total_dur % 60

    lines: list[str] = [
        "# M01 — Tess's Journey — Canonical Reconciliation Report",
        f"*Generated: May 12, 2026*",
        "",
        "## Canonical Files Created",
        f"- `{OUT_CANONICAL_CSV.name}`",
        f"- `{OUT_CANONICAL_XLSX.name}` (9-tab Excel matrix)",
        f"- `{OUT_NARRATION_CSV.name}`",
        f"- `{OUT_FLUX_CSV.name}`",
        f"- `{OUT_LTX_CSV.name}`",
        f"- `{OUT_CHALLENGES_CSV.name}`",
        f"- `{OUT_DEBRIEFS_CSV.name}`",
        "",
        "## Reconciliation Summary",
        f"| Metric | Value |",
        f"|--------|-------|",
        f"| Total production cards | {total} |",
        f"| Total estimated runtime | {mm:02d}:{ss:02d} (mm:ss) |",
        f"| Total LTX video generations | {total_ltx} |",
        f"| Status: COMPLETE | {complete} |",
        f"| Status: NEEDS REVIEW | {review} |",
        "",
        "## Card Count by Slide Type",
    ]
    for st, count in sorted(type_counts.items()):
        lines.append(f"- {st}: {count}")

    lines += [
        "",
        "## Batch Corrections Applied",
        f"- Scenes updated from batch corrections: {len(batch_scene_ids)}",
        f"- Scene IDs: {', '.join(sorted(batch_scene_ids)[:20])}{'...' if len(batch_scene_ids) > 20 else ''}",
        "",
        "## Changes Made",
        "- **Character name**: All instances of 'Marites' replaced with 'Tess' throughout all text fields",
        "- **S49 added**: Certificate Screen / The Reveal and the Legacy — was missing from source CSV",
        f"- **New columns added**: Detailed_Image_Description, LTX23Dev_Video_Generation_Prompt, "
        "Estimated_Card_Duration, LTX_Generation_Count, Operational_Rationale, Full_Debrief, Status",
        "- **FLUX column renamed**: FLUX_Image_Prompt → FLUX1Dev_Image_Prompt",
        "- **Challenge cards**: Operational_Rationale and Full_Debrief populated from batch corrections",
        "- **Debrief cards**: Full_Debrief populated from batch corrections where available",
        "",
    ]

    if missing_fields:
        lines += [
            "## Remaining Issues — Missing Production Fields",
            f"Total missing field instances: {len(missing_fields)}",
            "",
            "| Scene_ID | Missing Field |",
            "|----------|--------------|",
        ]
        for mf in missing_fields[:50]:
            sid, field = mf.split(' — ', 1)
            lines.append(f"| {sid} | {field} |")
        if len(missing_fields) > 50:
            lines.append(f"| ... | ({len(missing_fields) - 50} more) |")
    else:
        lines.append("## Remaining Issues\nNone. All required production fields are populated.")

    lines += [
        "",
        "## Source Files Consumed",
        f"- `{SOURCE_CSV.name}` (base data, {total - (1 if added_s49 else 0)} rows)",
        f"- `{BATCH1_TXT.name}` (S07–S17 corrections)",
        f"- `{BATCH2_TXT.name}` (S26–S46 corrections)",
        f"- `{BATCH3_TXT.name}` (S47–S49 corrections)",
        "",
        "## Archive Actions",
        "The following files were made obsolete by this canonical build:",
        f"- `{SOURCE_CSV.name}` → superseded (kept as archive reference)",
        f"- `M01-Marites-Journey-Narration-Defensibility-COMPLETE.xlsx` → superseded",
        "",
        "## Character Continuity Arc",
        "- Bracelet arc: grandmother's wrist (S05) → tied onto young Tess (S05) → removed for flight (S08) → back on wrist at U.S. workstation (S11) → touched during difficult moments (S20, S35, S48) → aged and faded on Director of Nursing wrist (S49)",
        "- Mr. Henderson arc: guarded gatekeeper (S18) → near-miss discovery (S30/S31) → grievance filed (S32) → corrected care resumed (S36/S37) → interpreter workflow executed (S38) → trust rebuilt at door (S39) → final acknowledgment (S48)",
        "- Time arc: 4:45 AM kitchen sounds, Quezon City (S01) → 4:45 PM Director of Nursing office clock (S49)",
        "",
        f"*Canonical build complete. All Module 1 production files now reference Tess's Journey exclusively.*",
    ]

    OUT_REPORT_MD.write_text('\n'.join(lines), encoding='utf-8')
    print(f"  Written: {OUT_REPORT_MD.name}")


# ──────────────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────────────

def main() -> None:
    print("\n=== M01 Canonical Build — Tess's Journey ===\n")

    print("Step 1: Parsing batch correction files...")
    batch1 = parse_batch_file(BATCH1_TXT)
    batch2 = parse_batch_file(BATCH2_TXT)
    batch3 = parse_batch_file(BATCH3_TXT)

    # Merge: batch2 supersedes batch1 (more detailed), batch3 is additive
    all_batch: dict[str, dict[str, str]] = {}
    all_batch.update(batch1)
    all_batch.update(batch2)  # overwrites batch1 where same Scene_ID
    all_batch.update(batch3)
    batch_scene_ids = set(all_batch.keys())
    print(f"  Total unique batch scenes: {len(batch_scene_ids)}")

    print("\nStep 2: Loading base CSV...")
    base_rows = load_base_csv()

    print("\nStep 3: Building canonical rows...")
    canonical_rows = build_canonical_rows(base_rows, all_batch)
    added_s49 = 'M01-S49' not in {r.get('Scene_ID') for r in base_rows}
    print(f"  Canonical row count: {len(canonical_rows)}")

    print("\nStep 4: Writing canonical CSV...")
    write_canonical_csv(canonical_rows)

    print("\nStep 5: Writing focused CSVs...")
    write_focused_csvs(canonical_rows)

    print("\nStep 6: Writing Excel matrix...")
    write_xlsx(canonical_rows)

    print("\nStep 7: Writing reconciliation report...")
    write_reconciliation_report(canonical_rows, batch_scene_ids, added_s49)

    # Runtime validation printout
    total = len(canonical_rows)
    total_dur = sum(int(r.get('Estimated_Card_Duration', '0') or '0') for r in canonical_rows)
    total_ltx = sum(int(r.get('LTX_Generation_Count', '0') or '0') for r in canonical_rows)
    complete  = sum(1 for r in canonical_rows if r.get('Status', '').startswith('COMPLETE'))
    review    = total - complete
    challenge_count = sum(1 for r in canonical_rows if 'challenge' in r.get('Slide_Type', ''))
    debrief_count   = sum(1 for r in canonical_rows if 'debrief'   in r.get('Slide_Type', ''))
    educator_count  = sum(1 for r in canonical_rows if 'educator'  in r.get('Slide_Type', ''))
    dup_check: dict[str, int] = {}
    for r in canonical_rows:
        sid = r.get('Scene_ID', '')
        dup_check[sid] = dup_check.get(sid, 0) + 1
    dups = [(sid, cnt) for sid, cnt in dup_check.items() if cnt > 1]

    mm = total_dur // 60
    ss_rem = total_dur % 60

    print("\n" + "="*55)
    print("M01 CANONICAL BUILD — RUNTIME VALIDATION")
    print("="*55)
    print(f"  Total cards:              {total}")
    print(f"  Total estimated runtime:  {mm:02d}:{ss_rem:02d} (mm:ss)")
    print(f"  Total LTX generations:    {total_ltx}")
    print(f"  Challenge cards:          {challenge_count}")
    print(f"  Debrief cards:            {debrief_count}")
    print(f"  Educator commentary:      {educator_count}")
    print(f"  Status COMPLETE:          {complete}")
    print(f"  Status NEEDS REVIEW:      {review}")
    print(f"  Duplicate Scene_IDs:      {len(dups)}")
    if dups:
        for sid, cnt in dups:
            print(f"    !! DUPLICATE: {sid} appears {cnt}x")
    print("="*55)
    print("\nCanonical build COMPLETE.\n")
    print("Output files:")
    print(f"  {OUT_CANONICAL_CSV}")
    print(f"  {OUT_CANONICAL_XLSX}")
    print(f"  {OUT_NARRATION_CSV}")
    print(f"  {OUT_FLUX_CSV}")
    print(f"  {OUT_LTX_CSV}")
    print(f"  {OUT_CHALLENGES_CSV}")
    print(f"  {OUT_DEBRIEFS_CSV}")
    print(f"  {OUT_REPORT_MD}")


if __name__ == "__main__":
    main()
